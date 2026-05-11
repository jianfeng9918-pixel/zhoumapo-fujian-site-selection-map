from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "data" / "preview-data.json"
REQUIRED_SHEETS = {"城市库", "商圈库", "点位库", "数据源登记表"}


def find_workbook() -> Path:
    candidates = []
    for path in Path("C:/CodexData").rglob("*.xlsx"):
        if ".bak-" in path.name:
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if REQUIRED_SHEETS.issubset(set(wb.sheetnames)):
                bonus = sum(1 for sheet in ["数据源等级表", "字段核验矩阵", "城市KPI表", "铺位线索表"] if sheet in wb.sheetnames)
                candidates.append((bonus, path.stat().st_mtime, path))
        except Exception:
            continue
    if candidates:
        return sorted(candidates, reverse=True)[0][2]
    raise FileNotFoundError("未找到包含城市库/商圈库/点位库的选址数据总表")


def value_to_text(value):
    if value is None:
        return "待补"
    text = str(value).strip()
    return text if text else "待补"


def sheet_to_records(wb, title):
    if title not in wb.sheetnames:
        return []
    ws = wb[title]
    headers = [value_to_text(cell.value) for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for header, value in zip(headers, row):
            item[header] = value_to_text(value)
        rows.append(item)
    return rows


def has_geo(row):
    return value_to_text(row.get("经度")) != "待补" and value_to_text(row.get("纬度")) != "待补"


def count_city(rows, city):
    return sum(1 for row in rows if row.get("城市") == city)


def avg_quality(rows):
    values = []
    for row in rows:
        try:
            values.append(float(row.get("数据质量评分", 0)))
        except Exception:
            pass
    return round(sum(values) / len(values), 1) if values else 0


def text(value):
    return value_to_text(value)


def to_number(value, default=0):
    try:
        parsed = float(str(value).replace("%", "").strip())
        return parsed if parsed == parsed else default
    except Exception:
        return default


def priority_rank(value):
    return {"P1": 0, "P2": 1, "P3": 2}.get(text(value), 9)


def decision_direct_recommended(row):
    return text(row.get("严格决策结论") or row.get("推荐结论") or row.get("决策结论")) == "推荐"


def decision_potential(row):
    return "潜力推荐" in text(row.get("严格决策结论") or row.get("推荐结论") or row.get("决策结论"))


def build_quality_dashboard(data):
    field_rows = data.get("fieldCoverage", [])
    core_rows = (
        data.get("cities", [])
        + data.get("districts", [])
        + data.get("businessAreas", [])
        + data.get("streets", [])
        + data.get("poi", [])
        + data.get("competitorStores", [])
        + data.get("rentSamples", [])
        + data.get("candidates", [])
        + data.get("siteLeads", [])
    )
    source_counts = Counter(text(row.get("来源等级")) for row in core_rows if text(row.get("来源等级")) != "待补")
    missing_rows = [row for row in field_rows if text(row.get("是否缺失")) == "是"]
    missing_by_field = Counter(text(row.get("字段")) for row in missing_rows)
    missing_by_city = Counter(text(row.get("城市")) for row in missing_rows if text(row.get("城市")) != "待补")
    low_quality = sorted(
        [
            row
            for row in data.get("businessAreas", []) + data.get("candidates", []) + data.get("siteLeads", [])
            if to_number(row.get("数据质量评分"), 0) < 70
        ],
        key=lambda row: (-to_number(row.get("综合评分"), 0), text(row.get("城市")), text(row.get("名称") or row.get("铺位线索名称"))),
    )[:60]
    city_quality = []
    for row in data.get("cityKpi", []):
        city_quality.append(
            {
                "城市": text(row.get("城市")),
                "推荐等级": text(row.get("推荐等级")),
                "平均数据质量分": text(row.get("平均数据质量分")),
                "商圈数": text(row.get("商圈数")),
                "核验任务数": text(row.get("核验任务数")),
                "P1任务数": text(row.get("P1任务数")),
                "缺竞品": text(row.get("缺竞品")),
                "缺租金": text(row.get("缺租金")),
                "缺照片": text(row.get("缺照片")),
                "缺平台数据": text(row.get("缺平台数据")),
                "下一步动作": text(row.get("下一步动作")),
            }
        )
    return {
        "sourceLevelCounts": dict(source_counts),
        "missingFieldCount": len(missing_rows),
        "missingByField": dict(missing_by_field.most_common(20)),
        "missingByCity": dict(missing_by_city.most_common()),
        "lowQualityObjects": low_quality,
        "cityQuality": city_quality,
        "qualityRule": "数据质量分低于70时，页面只显示潜力推荐/待核验，不显示直接推荐。",
    }


def build_map_layer_metrics(data):
    metrics = {}
    for city in [row for row in data["cities"] if row.get("省份") == "福建"]:
        city_name = text(city.get("城市")).replace("市", "")
        city_areas = [row for row in data["businessAreas"] if text(row.get("城市")).replace("市", "") == city_name]
        city_competitors = [row for row in data["competitorStores"] if text(row.get("城市")).replace("市", "") == city_name]
        city_tasks = [row for row in data["verificationTasks"] if text(row.get("城市")).replace("市", "") == city_name and text(row.get("状态")) != "已完成"]
        city_candidates = [row for row in data["candidates"] if text(row.get("城市")).replace("市", "") == city_name]
        field_missing = [
            row
            for row in data.get("fieldCoverage", [])
            if text(row.get("城市")).replace("市", "") == city_name and text(row.get("是否缺失")) == "是"
        ]
        qualities = [to_number(row.get("数据质量评分"), 0) for row in city_areas if to_number(row.get("数据质量评分"), 0)]
        metrics[city_name] = {
            "priorityScore": to_number(city.get("综合评分"), 0),
            "populationText": text(city.get("常住人口")),
            "economyText": text(city.get("GDP/社零/第三产业")),
            "areaCount": len(city_areas),
            "competitorCount": len(city_competitors),
            "taskCount": len(city_tasks),
            "p1TaskCount": sum(1 for row in city_tasks if text(row.get("优先级")) == "P1"),
            "avgQuality": round(sum(qualities) / max(1, len(qualities)), 1) if qualities else to_number(city.get("数据质量评分"), 0),
            "directRecommended": sum(1 for row in city_candidates if decision_direct_recommended(row)),
            "potentialRecommended": sum(1 for row in city_candidates if decision_potential(row)),
            "missingFieldCount": len(field_missing),
            "sourceLevel": text(city.get("来源等级")),
        }
    return metrics


def build_verification_queue(data):
    rows = [row for row in data.get("verificationTasks", []) if text(row.get("状态")) != "已完成"]
    rows = sorted(
        rows,
        key=lambda row: (
            priority_rank(row.get("优先级")),
            text(row.get("截止建议")),
            -to_number(row.get("数据质量评分"), 0),
            text(row.get("城市")),
        ),
    )
    return rows


FUJIAN_CITIES = {"福州", "厦门", "泉州", "漳州", "莆田", "三明", "南平", "龙岩", "宁德"}
CORE_COMPETITORS = {"小叫天", "醉得意", "四方桌", "大丰收", "姑奶奶"}

PUBLIC_STREET_COORDS = {
    ("福州", "仓山区", "金山街道"): (119.270, 26.047),
    ("福州", "晋安区", "岳峰镇"): (119.331, 26.088),
    ("福州", "鼓楼区", "杨桥路"): (119.280, 26.086),
    ("福州", "台江区", "工业路"): (119.303, 26.064),
    ("福州", "闽侯县", "上街镇"): (119.184, 26.063),
    ("福州", "台江区", "鳌峰街道/金融街片区待核验"): (119.344, 26.054),
    ("福州", "鼓楼区", "东街街道/南街街道待核验"): (119.300, 26.085),
    ("厦门", "思明区", "嘉禾路"): (118.122, 24.489),
    ("厦门", "思明区", "梧村街道"): (118.118, 24.468),
    ("厦门", "湖里区", "金山街道"): (118.173, 24.512),
    ("厦门", "思明区", "中山路"): (118.083, 24.454),
    ("厦门", "思明区", "鹭江街道"): (118.074, 24.457),
    ("厦门", "思明区", "中华街道/开元街道待核验"): (118.086, 24.455),
    ("泉州", "丰泽区", "泉秀街道/浦西片区待核验"): (118.606, 24.887),
}

PUBLIC_STORE_COORDS = {
    "福州永泰华煦中央府店": (118.941, 25.872),
    "福州晋安西园店": (119.324, 26.123),
    "福州晋安桂山店": (119.329, 26.113),
    "福州马尾儒江店": (119.425, 26.011),
    "南平武夷山宝龙店": (118.035, 27.755),
    "南平延平解放路店": (118.178, 26.641),
    "福州罗源凤蝶店": (119.552, 26.489),
    "泉州晋江英林万业城店": (118.591, 24.644),
    "泉州晋江磁灶店": (118.502, 24.724),
    "莆田正荣时代广场店": (119.021, 25.438),
}

PUBLIC_STREET_SEEDS = {
    ("福州", "鼓楼区", "东街街道"): ["东街口", "三坊七巷", "地铁东街口", "核心商圈"],
    ("福州", "鼓楼区", "南街街道"): ["南门兜", "三坊七巷", "东百中心", "文旅客流"],
    ("福州", "仓山区", "金山街道"): ["仓山万达", "金山住宅区", "浦上大道", "家庭客群"],
    ("福州", "晋安区", "岳峰镇"): ["东二环泰禾", "世欧广场", "福新路", "社区办公混合"],
    ("福州", "鼓楼区", "杨桥路"): ["福州万象城", "杨桥路", "西湖公园", "成熟住宅"],
    ("福州", "台江区", "工业路"): ["宝龙广场", "苏宁广场", "茶亭", "餐饮集聚"],
    ("福州", "闽侯县", "上街镇"): ["福州大学城", "学生客群", "上街商圈", "年轻消费"],
    ("厦门", "思明区", "嘉禾路"): ["SM城市广场", "地铁吕厝", "成熟商业轴", "办公住宅"],
    ("厦门", "思明区", "梧村街道"): ["厦门万象城", "厦门火车站", "罗宾森广场", "交通客流"],
    ("厦门", "湖里区", "金山街道"): ["湖里万达", "五缘湾", "湖里住宅", "家庭客群"],
    ("厦门", "思明区", "中山路"): ["中山路步行街", "中华城", "文旅客流", "老城商业"],
    ("厦门", "思明区", "鹭江街道"): ["鹭江道", "轮渡", "中华城", "游客客流"],
    ("泉州", "丰泽区", "泉秀街道/浦西片区待核验"): ["浦西万达", "领SHOW天地", "泉秀街道", "餐饮夜间消费"],
}


def norm_city(value):
    return text(value).replace("市", "")


def norm_district(value):
    return text(value).replace(" ", "")


def is_fujian_city(row):
    return norm_city(row.get("城市")) in FUJIAN_CITIES or text(row.get("省份")) == "福建"


def has_point(row):
    return text(row.get("经度")) != "待补" and text(row.get("纬度")) != "待补"


def point_tuple(row):
    if not has_point(row):
        return None
    return (to_number(row.get("经度")), to_number(row.get("纬度")))


def haversine_km(a, b):
    import math

    if not a or not b:
        return 999
    lon1, lat1 = a
    lon2, lat2 = b
    radius = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)
    h = math.sin(d_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    return 2 * radius * math.atan2(math.sqrt(h), math.sqrt(1 - h))


def is_own_store(row):
    return text(row.get("品牌")) == "周麻婆"


def core_competitor_brand(row):
    name = text(row.get("竞品品牌") or row.get("品牌") or row.get("门店名称"))
    for brand in CORE_COMPETITORS:
        if brand in name:
            return brand
    return ""


def decision_sort_key(row):
    grades = {"A": 4, "B": 3, "C": 2, "D": 1}
    return (grades.get(text(row.get("推荐等级")), 0), to_number(row.get("综合评分") or row.get("街道评分"), 0))


def known_districts_by_city(data):
    found = defaultdict(set)
    for sheet in ["districtMapMetrics", "districts", "businessAreas", "streetMapPoints", "storeDistribution"]:
        for row in data.get(sheet, []):
            city = norm_city(row.get("城市"))
            district = text(row.get("区县"))
            if city in FUJIAN_CITIES and district not in {"待补", city}:
                found[city].add(district)
    return {city: sorted(items, key=len, reverse=True) for city, items in found.items()}


def normalize_decision_source_data(data):
    districts = known_districts_by_city(data)
    for row in data.get("storeDistribution", []):
        city = norm_city(row.get("城市"))
        if city in FUJIAN_CITIES:
            district = text(row.get("区县"))
            haystack = " ".join(text(row.get(field)) for field in ["门店名称", "地址", "街道/片区"])
            if district in {"待补", city}:
                for candidate in districts.get(city, []):
                    if candidate in haystack:
                        row["区县"] = candidate
                        row["区县校正来源"] = "地址文本识别"
                        break
            if not has_point(row):
                name = text(row.get("门店名称"))
                for key, coords in PUBLIC_STORE_COORDS.items():
                    if key in name:
                        row["经度"], row["纬度"] = coords
                        row["坐标来源"] = "公开地址推定"
                        break
    for row in data.get("streetMapPoints", []):
        key = (norm_city(row.get("城市")), text(row.get("区县")), text(row.get("街道/片区")))
        if not has_point(row) and key in PUBLIC_STREET_COORDS:
            row["经度"], row["纬度"] = PUBLIC_STREET_COORDS[key]
            row["定位状态"] = "已定位"
            row["定位来源"] = "公开地名/关联商圈推定"
        seed = PUBLIC_STREET_SEEDS.get(key)
        if seed:
            row["公开POI线索"] = "、".join(seed)
            row["公开POI线索数"] = len(seed)
            row["公开资料状态"] = "公开地名线索"


def street_decision_label(row, own_near, core_near):
    score = to_number(row.get("街道评分") or row.get("综合评分"), 0)
    tag = text(row.get("机会标签"))
    own = own_near if own_near is not None else to_number(row.get("周麻婆现有门店数"), 0)
    core = core_near if core_near is not None else to_number(row.get("五大竞品门店数"), 0)
    if "过密" in tag or "谨慎" in tag or own >= 3:
        return "谨慎"
    if score >= 84 and own == 0 and core >= 2:
        return "优先进入"
    if score >= 78 and own <= 1:
        return "可观察"
    if own > 0:
        return "本店已覆盖"
    return "可观察"


def street_decision_reason(row, own_near, core_near, poi_count):
    label = street_decision_label(row, own_near, core_near)
    store_type = text(row.get("适合店型"))
    if label == "优先进入":
        return f"街道评分较高，周边本店覆盖弱，五大竞品已有验证，适合优先按{store_type}看铺。"
    if label == "谨慎":
        return "已有本店或竞争密度偏高，适合先复核租金、门头动线和同街区 cannibalization 风险。"
    if label == "本店已覆盖":
        return "已有周麻婆覆盖，适合观察补点空间或优化现有门店周边外卖半径。"
    return f"具备一定街道潜力和公开POI线索（{poi_count}项），适合作为备选街道继续观察。"


def build_decision_map(data):
    normalize_decision_source_data(data)
    own_points = [
        row for row in data.get("storeDistribution", [])
        if is_fujian_city(row) and is_own_store(row) and has_point(row)
    ]
    core_points = [
        {**row, "核心竞品品牌": core_competitor_brand(row)}
        for row in data.get("competitorStores", [])
        if is_fujian_city(row) and has_point(row) and core_competitor_brand(row)
    ]
    other_points = [
        row for row in data.get("competitorStores", [])
        if is_fujian_city(row) and has_point(row) and not core_competitor_brand(row)
    ]
    street_points = [
        row for row in data.get("streetMapPoints", [])
        if is_fujian_city(row)
    ]
    area_rows = [row for row in data.get("businessAreas", []) if is_fujian_city(row)]

    city_summaries = []
    for city in sorted(FUJIAN_CITIES):
        city_row = next((row for row in data.get("cities", []) if norm_city(row.get("城市")) == city and text(row.get("省份")) == "福建"), {})
        city_streets = [row for row in street_points if norm_city(row.get("城市")) == city]
        city_own = [row for row in own_points if norm_city(row.get("城市")) == city]
        city_core = [row for row in core_points if norm_city(row.get("城市")) == city]
        city_areas = [row for row in area_rows if norm_city(row.get("城市")) == city]
        city_summaries.append({
            "城市": city,
            "推荐等级": text(city_row.get("推荐等级")),
            "综合评分": text(city_row.get("综合评分")),
            "数据质量评分": text(city_row.get("数据质量评分")),
            "商圈样本数": len(city_areas),
            "街道样本数": len(city_streets),
            "周麻婆门店数": len(city_own),
            "五大竞品门店数": len(city_core),
            "高潜街道数": sum(1 for row in city_streets if "空白" in text(row.get("机会标签")) or "高潜" in text(row.get("机会标签"))),
        })

    district_summaries = []
    for row in data.get("districtMapMetrics", []):
        if not is_fujian_city(row):
            continue
        city = norm_city(row.get("城市"))
        district = norm_district(row.get("区县"))
        district_summaries.append({
            **row,
            "商圈样本数": len([item for item in area_rows if norm_city(item.get("城市")) == city and norm_district(item.get("区县")) == district]),
            "周麻婆门店数": len([item for item in own_points if norm_city(item.get("城市")) == city and norm_district(item.get("区县")) == district]),
            "五大竞品门店数": len([item for item in core_points if norm_city(item.get("城市")) == city and norm_district(item.get("区县")) == district]),
        })

    street_summaries = []
    street_decision_score = []
    for row in street_points:
        merged = {**row}
        city = norm_city(row.get("城市"))
        district = norm_district(row.get("区县"))
        street_text = text(row.get("街道/片区"))
        street_point = point_tuple(row)
        related_own = [
            item for item in own_points
            if norm_city(item.get("城市")) == city and (norm_district(item.get("区县")) == district or street_text in text(item.get("街道/片区")))
        ]
        related_core = [
            item for item in core_points
            if norm_city(item.get("城市")) == city and (norm_district(item.get("区县")) == district or street_text in text(item.get("街道/片区")))
        ]
        own_15 = len([item for item in own_points if norm_city(item.get("城市")) == city and haversine_km(street_point, point_tuple(item)) <= 1.5])
        core_30 = len([item for item in core_points if norm_city(item.get("城市")) == city and haversine_km(street_point, point_tuple(item)) <= 3])
        commercial_near = len([item for item in area_rows if norm_city(item.get("城市")) == city and (norm_district(item.get("区县")) == district or haversine_km(street_point, point_tuple(item)) <= 1.5)])
        public_poi_count = int(to_number(merged.get("公开POI线索数"), 0))
        poi_support = max(public_poi_count, commercial_near)
        decision_label = street_decision_label(merged, own_15, core_30)
        merged["周麻婆关联点位数"] = len(related_own)
        merged["五大竞品关联点位数"] = len(related_core)
        merged["周边1.5公里本店数"] = own_15
        merged["周边3公里五大竞品数"] = core_30
        merged["商圈/商业体线索数"] = commercial_near
        merged["POI支撑数"] = poi_support
        merged["街道潜力判断"] = decision_label
        merged["街道判断理由"] = street_decision_reason(merged, own_15, core_30, poi_support)
        merged["下一步动作"] = text(merged.get("下一步核验动作") or merged.get("下一步动作") or "复核平台表现、租金报价和现场门头动线")
        street_summaries.append(merged)
        street_decision_score.append(
            {
                "城市": city,
                "区县": text(row.get("区县")),
                "街道/片区": street_text,
                "经度": text(row.get("经度")),
                "纬度": text(row.get("纬度")),
                "推荐等级": text(row.get("推荐等级")),
                "综合评分": text(row.get("街道评分") or row.get("综合评分")),
                "街道评分": text(row.get("街道评分") or row.get("综合评分")),
                "数据质量评分": text(row.get("数据质量评分")),
                "街道潜力判断": decision_label,
                "适合店型": text(row.get("适合店型")),
                "机会标签": text(row.get("机会标签")),
                "周边1.5公里本店数": own_15,
                "周边3公里五大竞品数": core_30,
                "商圈/商业体线索数": commercial_near,
                "POI支撑数": poi_support,
                "公开POI线索": text(row.get("公开POI线索")),
                "街道判断理由": merged["街道判断理由"],
                "主要风险": text(row.get("主要风险") or row.get("竞品压力") or row.get("租金压力")),
                "下一步动作": merged["下一步动作"],
            }
        )

    return {
        "citySummaries": sorted(city_summaries, key=decision_sort_key, reverse=True),
        "districtSummaries": sorted(district_summaries, key=decision_sort_key, reverse=True),
        "streetSummaries": sorted(street_summaries, key=decision_sort_key, reverse=True),
        "streetDecisionScore": sorted(street_decision_score, key=decision_sort_key, reverse=True),
        "ownStorePoints": own_points,
        "coreCompetitorPoints": core_points,
        "otherCompetitorPoints": other_points,
    }


def build_stats(data):
    fujian_cities = [row for row in data["cities"] if row.get("省份") == "福建"]
    key_cities = ["福州", "厦门", "泉州"]
    line_rows = data["poi"] + data["competitorStores"] + data["rentSamples"]
    geo_rows = (
        data["cities"]
        + data["districts"]
        + data["businessAreas"]
        + data["streets"]
        + data["sites"]
        + data["poi"]
        + data["competitorStores"]
        + data["rentSamples"]
        + data.get("streetMapPoints", [])
    )
    geo_count = sum(1 for row in geo_rows if has_geo(row))
    city_depth = {
        city: {
            "districts": count_city(data["districts"], city),
            "areas": count_city(data["businessAreas"], city),
            "streets": count_city(data["streets"], city),
            "poi": count_city(data["poi"], city),
            "competitors": count_city(data["competitorStores"], city),
            "rent": count_city(data["rentSamples"], city),
            "radius": count_city(data["radiusStats"], city),
            "tasks": count_city(data["verificationTasks"], city),
            "candidates": count_city(data["candidates"], city),
            "lineTotal": count_city(line_rows, city),
        }
        for city in key_cities
    }
    quality_rows = data["businessAreas"] + data["districts"] + data["streets"]
    street_decisions = data.get("streetDecisions", [])
    street_map_points = data.get("streetMapPoints", [])
    district_map_metrics = data.get("districtMapMetrics", [])
    task_pending = sum(1 for row in data["verificationTasks"] if row.get("状态") == "待处理")
    candidate_recommended = sum(1 for row in data["candidates"] if decision_direct_recommended(row))
    candidate_potential = sum(1 for row in data["candidates"] if decision_potential(row))
    quality_blocked = sum(
        1
        for row in data["candidates"] + data.get("siteLeads", [])
        if text(row.get("推荐结论") or row.get("严格决策结论")) == "推荐" and to_number(row.get("数据质量评分"), 0) < 70
    )
    p1_tasks = sum(1 for row in data["verificationTasks"] if row.get("优先级") == "P1" and row.get("状态") != "已完成")
    return {
        "cityCount": len(data["cities"]),
        "fujianCityCount": len(fujian_cities),
        "districtCount": len(data["districts"]),
        "businessAreaCount": len(data["businessAreas"]),
        "streetCount": len(data["streets"]),
        "siteCount": len(data["sites"]),
        "poiCount": len(data["poi"]),
        "competitorStoreCount": len(data["competitorStores"]),
        "rentSampleCount": len(data["rentSamples"]),
        "visitTaskCount": len(data["visits"]),
        "imageTaskCount": len(data["images"]),
        "sourceCount": len(data["sources"]),
        "scoreCount": len(data["scores"]),
        "radiusCount": len(data["radiusStats"]),
        "verificationTaskCount": len(data["verificationTasks"]),
        "candidateCount": len(data["candidates"]),
        "reportCount": len(data["reports"]),
        "pendingTaskCount": task_pending,
        "p1TaskCount": p1_tasks,
        "recommendedCandidateCount": candidate_recommended,
        "potentialCandidateCount": candidate_potential,
        "qualityBlockedRecommendationCount": quality_blocked,
        "siteLeadCount": len(data.get("siteLeads", [])),
        "streetDecisionCount": len(street_decisions),
        "districtMapMetricCount": len(district_map_metrics),
        "streetMapPointCount": len(street_map_points),
        "streetMapPointLocatedCount": sum(1 for row in street_map_points if row.get("定位状态") == "已定位"),
        "streetMapPointPendingLocationCount": sum(1 for row in street_map_points if row.get("定位状态") != "已定位"),
        "storeDistributionCount": len(data.get("storeDistribution", [])),
        "competitorBrandCount": len(data.get("competitorBrands", [])),
        "requirementBacklogCount": len(data.get("requirementBacklog", [])),
        "fujianStreetDecisionCount": sum(1 for row in street_decisions if row.get("省份") == "福建"),
        "fieldCoverageCount": len(data.get("fieldCoverage", [])),
        "missingFieldCount": sum(1 for row in data.get("fieldCoverage", []) if row.get("是否缺失") == "是"),
        "avgQuality": avg_quality(quality_rows),
        "geoCount": geo_count,
        "geoTotal": len(geo_rows),
        "geoCoverage": round(geo_count / max(1, len(geo_rows)) * 100, 1),
        "cityDepth": city_depth,
    }


def main():
    workbook = find_workbook()
    wb = load_workbook(workbook, read_only=True, data_only=True)
    data = {
        "meta": {
            "title": "周麻婆福建选址决策工作台",
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "sourceWorkbook": str(workbook),
            "version": "V0.4 数据核验 + 决策交互版",
            "notes": "半径为现有经纬度近似计算；平台数据、租金、照片和地图API仍需持续核验。",
        },
        "cities": sheet_to_records(wb, "城市库"),
        "businessAreas": sheet_to_records(wb, "商圈库"),
        "sites": sheet_to_records(wb, "点位库"),
        "legacyCompetitors": sheet_to_records(wb, "竞品库"),
        "mallProjects": sheet_to_records(wb, "商场项目库"),
        "sources": sheet_to_records(wb, "数据源登记表"),
        "streets": sheet_to_records(wb, "街道库"),
        "scores": sheet_to_records(wb, "评分明细表"),
        "mapConfig": sheet_to_records(wb, "地图配置表"),
        "districts": sheet_to_records(wb, "区县库"),
        "poi": sheet_to_records(wb, "POI库"),
        "rentSamples": sheet_to_records(wb, "租金样本库"),
        "competitorStores": sheet_to_records(wb, "竞品门店库"),
        "visits": sheet_to_records(wb, "实地走访库"),
        "images": sheet_to_records(wb, "图片素材库"),
        "radiusStats": sheet_to_records(wb, "半径统计表"),
        "verificationTasks": sheet_to_records(wb, "核验任务表"),
        "candidates": sheet_to_records(wb, "候选清单表"),
        "reports": sheet_to_records(wb, "报告输出表"),
        "storeTypeWeights": sheet_to_records(wb, "店型权重表"),
        "sourceLevels": sheet_to_records(wb, "数据源等级表"),
        "fieldCoverage": sheet_to_records(wb, "字段核验矩阵"),
        "cityKpi": sheet_to_records(wb, "城市KPI表"),
        "districtKpi": sheet_to_records(wb, "区县KPI表"),
        "businessProfiles": sheet_to_records(wb, "商圈画像表"),
        "streetEnvironment": sheet_to_records(wb, "街道环境表"),
        "siteLeads": sheet_to_records(wb, "铺位线索表"),
        "streetDecisions": sheet_to_records(wb, "街道决策表"),
        "storeDistribution": sheet_to_records(wb, "门店分布表"),
        "competitorBrands": sheet_to_records(wb, "竞品品牌表"),
        "requirementBacklog": sheet_to_records(wb, "需求开发库"),
        "districtMapMetrics": sheet_to_records(wb, "区县地图指标"),
        "streetMapPoints": sheet_to_records(wb, "街道地图点"),
    }
    data["meta"]["version"] = "1.0 福建选址决策地图"
    data["meta"]["notes"] = "1.0聚焦福建选址决策主线：省级城市优先级、城市区县放大、区县街道机会点和街道对比分析。"
    data["qualityDashboard"] = build_quality_dashboard(data)
    data["mapLayerMetrics"] = build_map_layer_metrics(data)
    data["verificationQueue"] = build_verification_queue(data)
    data["decisionMap"] = build_decision_map(data)
    data["stats"] = build_stats(data)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUT)


if __name__ == "__main__":
    main()
