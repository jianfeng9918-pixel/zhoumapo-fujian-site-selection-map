from __future__ import annotations

import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PREVIEW_ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "01-选址数据总表" / "周麻婆选址数据总表.xlsx"
RAW_DIR = ROOT / "99-导入原始资料"
BRAND_SOURCE = RAW_DIR / "brand_competitors.xlsx"
MATERIAL_SOURCE = RAW_DIR / "member_material_addresses.xls"
TMP_XLRD = Path("C:/CodexData/.tmp_pydeps_xlrd")

TODAY = datetime.now().strftime("%Y-%m-%d")

FUJIAN_CITIES = {"福州", "厦门", "泉州", "漳州", "莆田", "三明", "南平", "龙岩", "宁德"}
TARGET_COMPETITORS = {
    "小叫天": ["小叫天"],
    "醉得意": ["醉得意", "最得意"],
    "四方桌": ["四方桌"],
    "大丰收": ["大丰收"],
    "姑奶奶": ["姑奶奶"],
}
OWN_BRAND = "周麻婆"
BRAND_CODES = {
    "小叫天": "XJT",
    "醉得意": "ZDY",
    "四方桌": "SFZ",
    "大丰收": "DFS",
    "姑奶奶": "GNN",
    "周麻婆": "ZMP",
}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_key(value) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def to_float(value):
    text = clean(value)
    if not text or text in {"-", "待补"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def to_int(value):
    parsed = to_float(value)
    return int(parsed) if parsed is not None else None


def source_level_text(level: str) -> str:
    return level


def find_workbook() -> Path:
    if WORKBOOK.exists():
        return WORKBOOK
    candidates = [path for path in ROOT.rglob("*.xlsx") if ".bak-" not in path.name and path.name != "brand_competitors.xlsx"]
    if not candidates:
        raise FileNotFoundError("未找到周麻婆选址数据总表")
    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def row_dict(headers, row) -> dict[str, str]:
    return {clean(header): clean(value) for header, value in zip(headers, row)}


def get_sheet_headers(ws) -> list[str]:
    return [clean(cell.value) for cell in ws[1]]


def rows_by_id(ws, id_header: str) -> dict[str, int]:
    headers = get_sheet_headers(ws)
    if id_header not in headers:
        return {}
    idx = headers.index(id_header) + 1
    result = {}
    for row in range(2, ws.max_row + 1):
        value = clean(ws.cell(row, idx).value)
        if value:
            result[value] = row
    return result


def remove_placeholder_rows(ws, id_header: str, prefixes: tuple[str, ...]):
    headers = get_sheet_headers(ws)
    if id_header not in headers:
        return
    idx = headers.index(id_header) + 1
    for row_num in range(ws.max_row, 1, -1):
        value = clean(ws.cell(row_num, idx).value)
        if any(value.startswith(prefix) for prefix in prefixes):
            ws.delete_rows(row_num, 1)


def upsert_row(ws, row_data: dict[str, str], id_header: str):
    headers = get_sheet_headers(ws)
    id_value = row_data[id_header]
    id_index = rows_by_id(ws, id_header)
    row_num = id_index.get(id_value, ws.max_row + 1)
    for col_num, header in enumerate(headers, start=1):
        if header in row_data:
            ws.cell(row_num, col_num, row_data[header])


def ensure_raw_files():
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    ascii_dir = Path("C:/CodexData/zhoumapo_import")
    if not BRAND_SOURCE.exists() and (ascii_dir / "brand_competitors.xlsx").exists():
        shutil.copy2(ascii_dir / "brand_competitors.xlsx", BRAND_SOURCE)
    if not MATERIAL_SOURCE.exists() and (ascii_dir / "member_material_addresses.xls").exists():
        shutil.copy2(ascii_dir / "member_material_addresses.xls", MATERIAL_SOURCE)
    if not BRAND_SOURCE.exists():
        raise FileNotFoundError(f"缺少导入源：{BRAND_SOURCE}")


def detect_brand(name: str) -> str | None:
    for brand, aliases in TARGET_COMPETITORS.items():
        if any(alias in name for alias in aliases):
            return brand
    if OWN_BRAND in name:
        return OWN_BRAND
    return None


def store_short_name(name: str) -> str:
    text = clean(name)
    match = re.search(r"[（(]([^）)]+)[）)]", text)
    if match:
        return match.group(1).strip()
    return text.replace("周麻婆·川式小炒", "").replace("周麻婆", "").strip(" ·-")


def quality_score(row: dict[str, str]) -> tuple[float, str, str]:
    score = 62
    if row.get("lng") and row.get("lat"):
        score += 12
    if row.get("评分"):
        score += 6
    if row.get("评论总数"):
        score += 5
    if row.get("人均消费"):
        score += 4
    if row.get("榜单信息"):
        score += 3
    if row.get("营业时间"):
        score += 2
    score = min(score, 88)
    if score >= 82:
        return score, "高", "字段较完整，仍需抽样复核营业状态和地址"
    if score >= 72:
        return score, "中高", "平台字段较完整，缺实地验证"
    return score, "中", "可用于竞争压力初判，需补平台截图/实地核验"


def field_completeness(row: dict[str, str], fields: list[str]) -> int:
    filled = sum(1 for field in fields if clean(row.get(field)))
    return round(filled / max(1, len(fields)) * 100)


def pressure_label(row: dict[str, str]) -> str:
    rating = to_float(row.get("评分"))
    comments = to_int(row.get("评论总数"))
    if rating and rating >= 4.5 and comments and comments >= 800:
        return "高"
    if comments and comments >= 300:
        return "中高"
    if rating and rating >= 4.2:
        return "中"
    return "待核验"


def read_brand_rows() -> list[dict[str, str]]:
    wb = load_workbook(BRAND_SOURCE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [clean(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = row_dict(headers, raw)
        city = row.get("城市", "").replace("市", "")
        if city not in FUJIAN_CITIES:
            continue
        row["城市"] = city
        brand = detect_brand(row.get("name", ""))
        if not brand:
            continue
        row["导入品牌"] = brand
        rows.append(row)
    return rows


def read_material_rows() -> list[dict[str, str]]:
    if not MATERIAL_SOURCE.exists():
        return []
    if TMP_XLRD.exists():
        sys.path.insert(0, str(TMP_XLRD))
    try:
        import xlrd
    except Exception:
        print("未安装 xlrd，已跳过旧版 xls 地址表。")
        return []

    book = xlrd.open_workbook(str(MATERIAL_SOURCE))
    sheet = book.sheet_by_index(0)
    rows = []
    for index in range(1, sheet.nrows):
        seq = clean(sheet.cell_value(index, 0))
        store_name = clean(sheet.cell_value(index, 1))
        city = clean(sheet.cell_value(index, 2)).replace("市", "")
        district = clean(sheet.cell_value(index, 3))
        address = clean(sheet.cell_value(index, 4))
        if not store_name or city not in FUJIAN_CITIES:
            continue
        rows.append(
            {
                "序号": seq,
                "门店": store_name,
                "城市": city,
                "地区": district,
                "地址": address,
            }
        )
    return rows


def competitor_row(source: dict[str, str]) -> dict[str, str]:
    brand = source["导入品牌"]
    code = BRAND_CODES[brand]
    shop_id = source.get("shopuuid") or source.get("shopid") or normalize_key(source.get("name"))
    quality, confidence, note = quality_score(source)
    completeness = field_completeness(
        source,
        ["name", "城市", "行政区", "regionName", "lng", "lat", "评分", "评论总数", "人均消费", "营业时间"],
    )
    category = " / ".join([value for value in [source.get("中类"), source.get("小类")] if value])
    rating_parts = [
        f"口味{source.get('评分详情_口味')}" if source.get("评分详情_口味") else "",
        f"环境{source.get('评分详情_环境')}" if source.get("评分详情_环境") else "",
        f"服务{source.get('评分详情_服务')}" if source.get("评分详情_服务") else "",
    ]
    return {
        "记录ID": f"CP-DP-{code}-{shop_id}",
        "省份": "福建",
        "城市": source.get("城市"),
        "区县": source.get("行政区"),
        "街道/片区": source.get("regionName") or "待补",
        "门店名称": source.get("name"),
        "竞品品牌": brand,
        "品类": category or "中式餐饮",
        "经度": source.get("lng") or "待补",
        "纬度": source.get("lat") or "待补",
        "关联商圈": source.get("regionName") or "待补",
        "距离判断": "待与候选街道/点位计算",
        "平台": "美团/大众点评",
        "评分": source.get("评分") or "待补",
        "评论数": source.get("评论总数") or "待补",
        "人均": source.get("人均消费") or "待补",
        "月销量/热度": source.get("榜单信息") or source.get("外卖") or source.get("团购") or "待补",
        "外卖表现": source.get("外卖") or "待补",
        "堂食表现": "；".join([part for part in rating_parts if part]) or source.get("评分详情") or "待补",
        "竞争压力": pressure_label(source),
        "数据置信度": confidence,
        "来源": "用户提供：品牌(1).xlsx",
        "来源类型": "平台数据导出/用户提供",
        "验证状态": "已导入-待抽样复核",
        "数据更新时间": TODAY,
        "备注": f"{note}；shopuuid={source.get('shopuuid')}",
        "数据质量评分": quality,
        "数据缺口": "缺实地照片/租金距离/营业状态抽样核验",
        "是否过期": "否",
        "是否冲突": "否",
        "核验动作": "抽样核验营业状态、门头、动线、与候选街道距离；补截图留档",
        "决策结论": "竞品核验线索",
        "店型匹配": "社区店/商圈店/购物中心店参考",
        "半径统计ID": "待补",
        "来源等级": source_level_text("L3"),
        "字段完整率": completeness,
        "可信度标签": confidence,
        "严格决策结论": "竞品核验线索",
        "推荐门槛说明": "竞品数据只用于竞争压力与商圈验证，不作为签约依据。",
        "V0.5核验状态": "已导入待抽样核验",
    }


def own_store_from_platform(source: dict[str, str]) -> dict[str, str]:
    shop_id = source.get("shopuuid") or source.get("shopid") or normalize_key(source.get("name"))
    return {
        "门店ID": f"OWN-DP-ZMP-{shop_id}",
        "品牌": "周麻婆",
        "门店名称": source.get("name"),
        "省份": "福建",
        "城市": source.get("城市"),
        "区县": source.get("行政区"),
        "街道/片区": source.get("regionName") or "待补",
        "地址": source.get("addressTextWq") or "待补",
        "经度": source.get("lng") or "待补",
        "纬度": source.get("lat") or "待补",
        "营业状态": source.get("经营状态") or "待补",
        "店型": "待识别",
        "来源": "用户提供：品牌(1).xlsx",
        "来源等级": "L3",
        "备注": f"平台门店数据；短名={store_short_name(source.get('name'))}",
        "数据更新时间": TODAY,
    }


def own_store_from_material(row: dict[str, str]) -> dict[str, str]:
    seq = row.get("序号") or normalize_key(row.get("门店"))
    return {
        "门店ID": f"OWN-MAT-ZMP-{seq}",
        "品牌": "周麻婆",
        "门店名称": row.get("门店"),
        "省份": "福建",
        "城市": row.get("城市"),
        "区县": row.get("地区"),
        "街道/片区": "待补",
        "地址": row.get("地址"),
        "经度": "待补",
        "纬度": "待补",
        "营业状态": "待核验",
        "店型": "待识别",
        "来源": "用户提供：地址加单号12月新会员物料搜集表-地址.xls",
        "来源等级": "L3",
        "备注": "内部地址表补充，待高德/百度地理编码补经纬度",
        "数据更新时间": TODAY,
    }


def brand_summary_rows(competitor_rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in competitor_rows:
        grouped[row["竞品品牌"]].append(row)
    result = {}
    for brand, rows in grouped.items():
        prices = [to_float(row.get("人均")) for row in rows]
        prices = [price for price in prices if price is not None]
        avg_price = round(mean(prices), 1) if prices else None
        city_counts = Counter(row["城市"] for row in rows)
        top_city_text = "、".join(f"{city}{count}家" for city, count in city_counts.most_common(3))
        result[brand] = {
            "门店数据状态": f"福建已导入{len(rows)}家；重点城市：{top_city_text}；待抽样核验",
            "价格带": f"人均约{avg_price}元（平台字段均值）" if avg_price else "30-70元待复核",
        }
    return result


def update_brand_table(ws, competitor_rows: list[dict[str, str]]):
    summaries = brand_summary_rows(competitor_rows)
    headers = get_sheet_headers(ws)
    name_col = headers.index("品牌名称") + 1
    existing = {}
    typo_row = None
    for row_num in range(2, ws.max_row + 1):
        name = clean(ws.cell(row_num, name_col).value)
        if name == "最得意":
            typo_row = row_num
            name = "醉得意"
            ws.cell(row_num, name_col, name)
        if name:
            existing[name] = row_num

    core_brands = ["小叫天", "醉得意", "四方桌", "大丰收", "姑奶奶"]
    base_info = {
        "小叫天": ("福建核心竞品", "川菜/家常菜", "福州及地市街边/商圈门店密度较高，需要判断贴身竞争。"),
        "醉得意": ("福建核心竞品", "家常菜/中式正餐", "福建省内门店量级高，是周麻婆商圈验证和竞品压力判断的关键品牌。"),
        "四方桌": ("福建核心竞品", "福建菜/客家土菜", "偏聚餐和区域家常菜，可观察购物中心与家庭消费场景。"),
        "大丰收": ("福建核心竞品", "脆鱼/干锅/中式正餐", "商场和区域型门店较多，可作为同价位餐饮集聚验证。"),
        "姑奶奶": ("福建核心竞品", "中式家常菜/区域连锁", "宁德、漳州等城市覆盖较明显，用于判断区域竞争格局。"),
    }
    for index, brand in enumerate(core_brands, start=1):
        row_num = existing.get(brand) or typo_row or ws.max_row + 1
        summary = summaries.get(brand, {})
        values = {
            "品牌ID": f"COMP-BRAND-{index:03d}",
            "品牌名称": brand,
            "竞品类型": base_info[brand][0],
            "餐饮业态": base_info[brand][1],
            "价格带": summary.get("价格带", "30-70元待复核"),
            "重点说明": base_info[brand][2],
            "是否核心竞品": "是",
            "门店数据状态": summary.get("门店数据状态", "本轮未匹配到门店，待补"),
            "来源": "用户提供：品牌(1).xlsx",
            "来源等级": "L3",
            "核验动作": "抽样核验营业状态、评分、销量、门头、动线与候选街道距离",
            "数据更新时间": TODAY,
        }
        for col_num, header in enumerate(headers, start=1):
            if header in values:
                ws.cell(row_num, col_num, values[header])
        existing[brand] = row_num
        typo_row = None

    for row_num in range(ws.max_row, 1, -1):
        name = clean(ws.cell(row_num, name_col).value)
        if name and name not in core_brands:
            ws.delete_rows(row_num, 1)


def update_sources_and_log(wb, counts: dict[str, int]):
    source_ws = wb["数据源登记表"]
    source_rows = [
        {
            "来源ID": "SRC-20260510-BRAND-COMPETITOR",
            "来源类型": "用户提供平台数据",
            "省份": "福建",
            "城市": "福建9城",
            "关联对象类型": "竞品门店/周麻婆门店",
            "关联对象名称": "小叫天、醉得意、四方桌、大丰收、姑奶奶、周麻婆",
            "资料标题": "品牌(1).xlsx",
            "资料日期": "2026-05-10",
            "采集日期": TODAY,
            "采集方式": "用户提供Excel",
            "文件/链接路径": str(BRAND_SOURCE),
            "数据字段": "门店名、城市、区县、商圈、地址、经纬度、评分、人均、评论、榜单、营业时间",
            "置信度影响": "L3，可用于竞品密度与商圈验证，仍需抽样核验",
            "可用结论": f"导入竞品{counts['competitors']}家、周麻婆平台门店{counts['own_platform']}家",
            "限制说明": "平台导出数据可能存在营业状态和评分时点变化，不能单独作为签约依据",
            "更新人": "Codex",
            "复核人": "待补",
            "备注": "仅导入福建城市",
            "登记日期": TODAY,
            "登记人": "Codex",
            "复核状态": "待抽样复核",
        },
        {
            "来源ID": "SRC-20260510-ZMP-MATERIAL-ADDRESS",
            "来源类型": "用户提供内部地址表",
            "省份": "福建",
            "城市": "福建相关城市",
            "关联对象类型": "周麻婆门店",
            "关联对象名称": "12月新会员物料地址表",
            "资料标题": "地址加单号12月新会员物料搜集表-地址.xls",
            "资料日期": "2026-05-10",
            "采集日期": TODAY,
            "采集方式": "用户提供Excel",
            "文件/链接路径": str(MATERIAL_SOURCE),
            "数据字段": "门店、城市、地区、地址",
            "置信度影响": "L3，补充周麻婆现有门店地址，但缺经纬度",
            "可用结论": f"地址表福建记录{counts['material_total']}条；新增未匹配地址{counts['own_material_added']}条",
            "限制说明": "缺经纬度和营业状态，需要高德/百度地理编码与人工复核",
            "更新人": "Codex",
            "复核人": "待补",
            "备注": "非福建记录已跳过",
            "登记日期": TODAY,
            "登记人": "Codex",
            "复核状态": "待地理编码",
        },
    ]
    for row in source_rows:
        upsert_row(source_ws, row, "来源ID")

    log_ws = wb["更新日志"]
    log_row = {
        "日志ID": "LOG-20260510-FJ-BRAND-IMPORT",
        "更新时间": TODAY,
        "更新类型": "数据导入",
        "更新对象类型": "竞品门店/门店分布/品牌表",
        "省份": "福建",
        "城市": "福建9城",
        "对象名称": "五个核心竞品与周麻婆门店",
        "更新前结论": "竞品门店仅少量线索，周麻婆门店图层为空模板",
        "更新后结论": f"导入竞品{counts['competitors']}家、周麻婆门店{counts['own_total']}条，地图可显示有经纬度门店",
        "变化原因": "用户提供完整品牌门店表和内部地址表",
        "更新人": "Codex",
        "复核人": "待补",
        "下一步动作": "按福州/厦门/泉州抽样核验营业状态、平台截图、门头照片；对无坐标地址做高德地理编码",
        "备注": "低质量或未核验数据仍不得作为签约依据",
    }
    upsert_row(log_ws, log_row, "日志ID")


def row_matches_decision(row: dict[str, str], decision: dict[str, str]) -> bool:
    if row.get("城市") != decision.get("城市") or row.get("区县") != decision.get("区县"):
        return False
    region = normalize_key(row.get("街道/片区") or row.get("关联商圈"))
    street = normalize_key(decision.get("街道/片区"))
    area = normalize_key(decision.get("关联商圈"))
    if not region:
        return False
    return bool(
        region == street
        or region == area
        or (street and (street in region or region in street))
        or (area and (area in region or region in area))
    )


def update_street_decisions(ws, own_rows: list[dict[str, str]], comp_rows: list[dict[str, str]]):
    headers = get_sheet_headers(ws)
    rows = []
    for row_num in range(2, ws.max_row + 1):
        item = {header: clean(ws.cell(row_num, col_num).value) for col_num, header in enumerate(headers, start=1)}
        rows.append((row_num, item))

    header_index = {header: index + 1 for index, header in enumerate(headers)}
    for row_num, decision in rows:
        own_count = sum(1 for row in own_rows if row_matches_decision(row, decision))
        comp_count = sum(1 for row in comp_rows if row_matches_decision(row, decision))
        district_comp_count = sum(1 for row in comp_rows if row.get("城市") == decision.get("城市") and row.get("区县") == decision.get("区县"))
        district_own_count = sum(1 for row in own_rows if row.get("城市") == decision.get("城市") and row.get("区县") == decision.get("区县"))
        if own_count or comp_count:
            ws.cell(row_num, header_index["周麻婆现有门店数"], own_count)
            ws.cell(row_num, header_index["友商/竞品门店数"], comp_count)
            basis = clean(decision.get("主要依据"))
            addition = f"平台导入显示本街道/商圈匹配周麻婆{own_count}家、核心竞品{comp_count}家"
            if addition not in basis:
                ws.cell(row_num, header_index["主要依据"], f"{basis}；{addition}" if basis else addition)
        elif district_comp_count or district_own_count:
            risk = clean(decision.get("主要风险"))
            addition = f"同区县平台导入周麻婆{district_own_count}家、核心竞品{district_comp_count}家，街道归属需进一步核验"
            if addition not in risk:
                ws.cell(row_num, header_index["主要风险"], f"{risk}；{addition}" if risk else addition)
        if comp_count >= 6 and "竞品压力" in header_index:
            ws.cell(row_num, header_index["竞品压力"], "高")
        elif comp_count >= 3 and "竞品压力" in header_index:
            ws.cell(row_num, header_index["竞品压力"], "中高")
        if "数据更新时间" in header_index:
            ws.cell(row_num, header_index["数据更新时间"], TODAY)


def update_kpi_counts(wb, own_rows: list[dict[str, str]], comp_rows: list[dict[str, str]]):
    def update_sheet(ws, key_fields: list[str]):
        headers = get_sheet_headers(ws)
        header_index = {header: index + 1 for index, header in enumerate(headers)}
        for row_num in range(2, ws.max_row + 1):
            key = {field: clean(ws.cell(row_num, header_index[field]).value) for field in key_fields}
            competitors = [
                row
                for row in comp_rows
                if all(clean(row.get(field)) == key[field] for field in key_fields if field in row)
            ]
            if competitors and "竞品数" in header_index:
                ws.cell(row_num, header_index["竞品数"], len(competitors))
            if competitors and "缺竞品" in header_index:
                ws.cell(row_num, header_index["缺竞品"], 0)
            if "下一步动作" in header_index:
                current = clean(ws.cell(row_num, header_index["下一步动作"]).value)
                addition = "已导入五个核心竞品门店，下一步抽样核验评分/营业状态/门头照片"
                if competitors and addition not in current:
                    ws.cell(row_num, header_index["下一步动作"], f"{current}；{addition}" if current else addition)
            if "数据更新时间" in header_index:
                ws.cell(row_num, header_index["数据更新时间"], TODAY)

    update_sheet(wb["城市KPI表"], ["城市"])
    update_sheet(wb["区县KPI表"], ["城市", "区县"])


def import_data():
    ensure_raw_files()
    workbook = find_workbook()
    backup = workbook.with_name(f"{workbook.stem}.bak-brand-import-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook.suffix}")
    shutil.copy2(workbook, backup)

    brand_rows = read_brand_rows()
    competitor_sources = [row for row in brand_rows if row["导入品牌"] in TARGET_COMPETITORS]
    own_sources = [row for row in brand_rows if row["导入品牌"] == OWN_BRAND]
    material_sources = read_material_rows()

    wb = load_workbook(workbook)
    comp_ws = wb["竞品门店库"]
    own_ws = wb["门店分布表"]
    brand_ws = wb["竞品品牌表"]
    remove_placeholder_rows(own_ws, "门店ID", ("STORE-TEMPLATE",))

    competitor_rows = [competitor_row(row) for row in competitor_sources]
    for row in competitor_rows:
        upsert_row(comp_ws, row, "记录ID")

    own_rows = [own_store_from_platform(row) for row in own_sources]
    platform_short_names = {normalize_key(store_short_name(row.get("门店名称"))) for row in own_rows}
    platform_addresses = {normalize_key(row.get("地址")) for row in own_rows if row.get("地址")}
    material_added = 0
    for row in material_sources:
        name_key = normalize_key(row.get("门店"))
        address_key = normalize_key(row.get("地址"))
        matched = name_key in platform_short_names or address_key in platform_addresses
        if not matched:
            own_rows.append(own_store_from_material(row))
            material_added += 1

    for row in own_rows:
        upsert_row(own_ws, row, "门店ID")

    update_brand_table(brand_ws, competitor_rows)
    update_street_decisions(wb["街道决策表"], own_rows, competitor_rows)
    update_kpi_counts(wb, own_rows, competitor_rows)
    counts = {
        "competitors": len(competitor_rows),
        "own_platform": len(own_sources),
        "own_material_added": material_added,
        "own_total": len(own_rows),
        "material_total": len(material_sources),
    }
    update_sources_and_log(wb, counts)

    wb.save(workbook)
    return workbook, backup, counts, Counter(row["竞品品牌"] for row in competitor_rows), Counter(row["城市"] for row in competitor_rows)


def main():
    workbook, backup, counts, brand_counts, city_counts = import_data()
    print(f"Workbook updated: {workbook}")
    print(f"Backup created: {backup}")
    print(f"Competitor stores imported: {counts['competitors']}")
    print(f"Own stores imported/displayed: {counts['own_total']} (platform={counts['own_platform']}, material added={counts['own_material_added']})")
    print("Brand counts:")
    for brand, count in brand_counts.most_common():
        print(f"  {brand}: {count}")
    print("City counts:")
    for city, count in city_counts.most_common():
        print(f"  {city}: {count}")


if __name__ == "__main__":
    main()
