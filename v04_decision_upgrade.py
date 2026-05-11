from __future__ import annotations

import math
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


TODAY = datetime.now().strftime("%Y-%m-%d")
ROOT = Path(__file__).resolve().parent

REQUIRED_SHEETS = {"城市库", "商圈库", "点位库", "数据源登记表"}

QUALITY_COLUMNS = ["数据质量评分", "数据缺口", "是否过期", "是否冲突", "核验动作", "决策结论", "店型匹配", "半径统计ID"]

STORE_TYPE_WEIGHTS = [
    ("社区店", "商圈/街道/候选环境", "住宅与社区密度", 28, "优先看常住人口、晚餐复购、社区底商可得性。"),
    ("社区店", "商圈/街道/候选环境", "租金回本压力", 24, "租金和转让费过高会直接降级。"),
    ("社区店", "商圈/街道/候选环境", "外卖3公里覆盖", 22, "外卖半径内住宅、学校、办公越多越好。"),
    ("社区店", "商圈/街道/候选环境", "竞品压力", 16, "竞品多但验证强可加分，过密且低价则降分。"),
    ("社区店", "商圈/街道/候选环境", "数据置信度", 10, "实地和平台数据越完整越高。"),
    ("商圈店", "商圈/街道/候选环境", "商圈客流和餐饮氛围", 30, "优先看购物、夜间、文旅和餐饮聚集。"),
    ("商圈店", "商圈/街道/候选环境", "堂食1-1.5公里", 24, "堂食客群和门头动线是核心。"),
    ("商圈店", "商圈/街道/候选环境", "品牌曝光价值", 18, "首店、样板店和核心商圈曝光加分。"),
    ("商圈店", "商圈/街道/候选环境", "租金压力", 18, "高租金必须有强客流和转化依据。"),
    ("商圈店", "商圈/街道/候选环境", "数据置信度", 10, "平台和实地数据越完整越高。"),
    ("购物中心店", "商圈/街道/候选环境", "商业体等级和主力店", 30, "重点看商场成熟度、主力店和餐饮楼层。"),
    ("购物中心店", "商圈/街道/候选环境", "品牌适配和曝光", 22, "品牌形象、招商位置和竞品组合决定价值。"),
    ("购物中心店", "商圈/街道/候选环境", "租金和合同条件", 24, "扣点、保底、物业费、装修期需核验。"),
    ("购物中心店", "商圈/街道/候选环境", "堂食转化", 14, "楼层动线和餐饮区聚集影响转化。"),
    ("购物中心店", "商圈/街道/候选环境", "数据置信度", 10, "招商资料和实地照片权重高。"),
    ("高校/办公店", "商圈/街道/候选环境", "工作日/学生客流", 30, "午餐、晚餐和外卖峰值明显加分。"),
    ("高校/办公店", "商圈/街道/候选环境", "外卖3公里覆盖", 26, "宿舍、办公楼、产业园覆盖越强越好。"),
    ("高校/办公店", "商圈/街道/候选环境", "价格带适配", 18, "价格敏感区域需控制租金和面积。"),
    ("高校/办公店", "商圈/街道/候选环境", "季节/周末波动", 14, "高校寒暑假和办公周末波动需扣分。"),
    ("高校/办公店", "商圈/街道/候选环境", "数据置信度", 12, "平台外卖和实地时段观察权重高。"),
]


def find_workbook() -> Path:
    for path in Path("C:/CodexData").rglob("*.xlsx"):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if REQUIRED_SHEETS.issubset(set(wb.sheetnames)):
                return path
        except Exception:
            continue
    raise FileNotFoundError("未找到周麻婆选址数据总表.xlsx")


def value(value):
    if value is None:
        return ""
    return str(value).strip()


def text(value, fallback="待补"):
    raw = value if isinstance(value, str) else ("" if value is None else str(value))
    raw = raw.strip()
    return raw if raw else fallback


def ensure_headers(ws, headers):
    existing = [value(cell.value) for cell in ws[1]]
    if not any(existing):
        ws.append(headers)
        existing = [value(cell.value) for cell in ws[1]]
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=len(existing) + 1).value = header
            existing.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="15724C")


def ensure_sheet(wb, title, headers):
    if title not in wb.sheetnames:
        ws = wb.create_sheet(title)
        ws.append(headers)
    else:
        ws = wb[title]
    ensure_headers(ws, headers)
    return ws


def header_map(ws):
    return {value(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}


def records(ws):
    headers = [value(cell.value) for cell in ws[1]]
    out = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value(cell) for cell in row):
            continue
        item = {"__row": idx}
        for header, cell in zip(headers, row):
            item[header] = text(cell)
        out.append(item)
    return out


def clear_generated(ws, prefixes):
    for row in range(ws.max_row, 1, -1):
        key = value(ws.cell(row=row, column=1).value)
        if any(key.startswith(prefix) for prefix in prefixes):
            ws.delete_rows(row)


def append_records(ws, rows):
    headers = header_map(ws)
    for row in rows:
        next_row = ws.max_row + 1
        for key, val in row.items():
            if key not in headers:
                ensure_headers(ws, [key])
                headers = header_map(ws)
            ws.cell(row=next_row, column=headers[key]).value = val


def as_float(value_):
    try:
        parsed = float(str(value_).strip())
        return parsed if math.isfinite(parsed) else None
    except Exception:
        return None


def has_geo(row):
    return as_float(row.get("经度")) is not None and as_float(row.get("纬度")) is not None


def haversine_km(a_lon, a_lat, b_lon, b_lat):
    radius = 6371.0
    lon1, lat1, lon2, lat2 = map(math.radians, [a_lon, a_lat, b_lon, b_lat])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * radius * math.asin(math.sqrt(h))


def score_value(row):
    return int(as_float(row.get("综合评分") or row.get("环境评分")) or 0)


def source_points(rows, city):
    points = []
    for row in rows:
        if row.get("城市") != city or not has_geo(row):
            continue
        points.append(row)
    return points


def classify_quality(row, radius_row=None):
    source_type = text(row.get("来源类型"), "")
    confidence = text(row.get("数据置信度") or row.get("验证状态"), "")
    gaps = []
    quality = 32
    if has_geo(row):
        quality += 18
    else:
        gaps.append("缺经纬度")
    if "官方" in source_type:
        quality += 20
    elif "公开" in source_type:
        quality += 14
    elif "AI" in source_type:
        quality += 5
        gaps.append("AI线索待核验")
    else:
        gaps.append("来源类型待补")
    if confidence == "高":
        quality += 16
    elif confidence == "中":
        quality += 11
    elif confidence == "低":
        quality += 5
    else:
        gaps.append("置信度待补")
    if radius_row:
        if int(radius_row.get("3公里竞品数", 0)) > 0:
            quality += 8
        else:
            gaps.append("缺竞品核验")
        if int(radius_row.get("3公里租金样本数", 0)) > 0:
            quality += 8
        else:
            gaps.append("缺租金样本")
        if int(radius_row.get("3公里POI数", 0)) > 0:
            quality += 6
        else:
            gaps.append("缺POI半径")
    else:
        gaps.append("缺半径统计")
    if text(row.get("主要风险"), "") in {"", "待补"}:
        gaps.append("缺风险描述")
    if text(row.get("下一步动作"), "") in {"", "待补"}:
        gaps.append("缺下一步动作")
    quality = max(0, min(100, quality))
    return quality, "；".join(dict.fromkeys(gaps)) or "暂无明显缺口"


def decision(score, quality):
    if score >= 84 and quality >= 68:
        return "推荐"
    if score >= 74 and quality >= 50:
        return "谨慎推荐"
    if quality < 45:
        return "先补数据"
    return "谨慎"


def store_fit(row):
    kind = text(row.get("商圈类型") or row.get("适合店型") or row.get("消费场景"), "")
    lower = kind
    fits = []
    if any(token in lower for token in ["社区", "住宅", "县域"]):
        fits.append("社区店")
    if any(token in lower for token in ["商圈", "购物", "文旅", "餐饮"]):
        fits.append("商圈店")
    if any(token in lower for token in ["购物中心", "万达", "商场", "商业体"]):
        fits.append("购物中心店")
    if any(token in lower for token in ["高校", "办公", "产业", "CBD", "园区"]):
        fits.append("高校/办公店")
    return "、".join(dict.fromkeys(fits)) or "商圈店"


def radius_judgement(radius_row):
    poi15 = int(radius_row.get("1.5公里POI数", 0))
    comp15 = int(radius_row.get("1.5公里竞品数", 0))
    poi3 = int(radius_row.get("3公里POI数", 0))
    comp3 = int(radius_row.get("3公里竞品数", 0))
    if poi15 >= 8 and comp15 >= 2:
        dine = "堂食圈强：1.5公里内客群和竞品验证较充分"
    elif poi15 >= 4:
        dine = "堂食圈中：具备客群线索，需实地验证门头动线"
    else:
        dine = "堂食圈弱：需补客流和门头照片"
    if poi3 >= 15 and comp3 >= 3:
        delivery = "外卖圈强：3公里内POI和竞品线索较多"
    elif poi3 >= 8:
        delivery = "外卖圈中：可做外卖半径验证"
    else:
        delivery = "外卖圈弱：需补平台热力和配送范围"
    return dine, delivery


def build_radius_rows(areas, poi_rows, competitor_rows, rent_rows):
    out = []
    for index, area in enumerate([r for r in areas if r.get("省份") == "福建"], start=1):
        lon = as_float(area.get("经度"))
        lat = as_float(area.get("纬度"))
        if lon is None or lat is None:
            out.append({
                "记录ID": f"RAD-V04-{index:04d}",
                "对象类型": "商圈",
                "对象ID": area.get("记录ID"),
                "省份": area.get("省份"),
                "城市": area.get("城市"),
                "区县": area.get("区县"),
                "名称": area.get("名称"),
                "经度": area.get("经度"),
                "纬度": area.get("纬度"),
                "数据质量评分": 30,
                "数据缺口": "缺经纬度，无法计算半径",
                "核验动作": "先补经纬度，再计算堂食/外卖半径",
                "数据更新时间": TODAY,
                "来源": "V0.4半径统计",
                "验证状态": "待补坐标",
            })
            continue
        city = area.get("城市")
        point_groups = {
            "POI": source_points(poi_rows, city),
            "竞品": source_points(competitor_rows, city),
            "租金": source_points(rent_rows, city),
        }
        distances = {}
        nearest_comp = ("待补", 999)
        nearest_rent = ("待补", 999)
        for group_name, group_rows in point_groups.items():
            distances[group_name] = []
            for point in group_rows:
                dist = haversine_km(lon, lat, as_float(point.get("经度")), as_float(point.get("纬度")))
                distances[group_name].append((dist, point))
                if group_name == "竞品" and dist < nearest_comp[1]:
                    nearest_comp = (text(point.get("门店名称") or point.get("竞品品牌")), dist)
                if group_name == "租金" and dist < nearest_rent[1]:
                    nearest_rent = (text(point.get("租金区间") or point.get("铺位类型")), dist)
        row = {
            "记录ID": f"RAD-V04-{index:04d}",
            "对象类型": "商圈",
            "对象ID": area.get("记录ID"),
            "省份": area.get("省份"),
            "城市": city,
            "区县": area.get("区县"),
            "名称": area.get("名称"),
            "经度": lon,
            "纬度": lat,
            "1公里POI数": sum(1 for d, _ in distances["POI"] if d <= 1),
            "1公里竞品数": sum(1 for d, _ in distances["竞品"] if d <= 1),
            "1公里租金样本数": sum(1 for d, _ in distances["租金"] if d <= 1),
            "1.5公里POI数": sum(1 for d, _ in distances["POI"] if d <= 1.5),
            "1.5公里竞品数": sum(1 for d, _ in distances["竞品"] if d <= 1.5),
            "1.5公里租金样本数": sum(1 for d, _ in distances["租金"] if d <= 1.5),
            "3公里POI数": sum(1 for d, _ in distances["POI"] if d <= 3),
            "3公里竞品数": sum(1 for d, _ in distances["竞品"] if d <= 3),
            "3公里租金样本数": sum(1 for d, _ in distances["租金"] if d <= 3),
            "最近竞品": nearest_comp[0],
            "最近竞品距离km": round(nearest_comp[1], 2) if nearest_comp[1] < 999 else "待补",
            "最近租金": nearest_rent[0],
            "数据更新时间": TODAY,
            "来源": "V0.4现有经纬度近似半径计算",
            "验证状态": "待地图API复核",
        }
        dine, delivery = radius_judgement(row)
        row["堂食半径判断"] = dine
        row["外卖半径判断"] = delivery
        quality, gaps = classify_quality(area, row)
        row["数据质量评分"] = quality
        row["数据缺口"] = gaps
        row["核验动作"] = "补平台竞品截图、租金报价、实地照片和高德/百度POI半径"
        out.append(row)
    return out


def build_tasks(areas, radius_rows, images):
    radius_by_id = {row["对象ID"]: row for row in radius_rows}
    image_names = {text(row.get("关联对象")) for row in images}
    tasks = []
    task_index = 1
    for area in areas:
        if area.get("省份") != "福建":
            continue
        radius = radius_by_id.get(area.get("记录ID"))
        quality, gaps = classify_quality(area, radius)
        priority = "P1" if area.get("城市") == "福州" and score_value(area) >= 80 else "P2" if score_value(area) >= 72 else "P3"
        gap_text = gaps if gaps != "暂无明显缺口" else "补平台数据和实地照片"
        task_types = []
        if "缺竞品" in gap_text or "AI线索" in gap_text:
            task_types.append(("竞品平台核验", "美团/点评评分、评论、人均、销量、榜单、距离截图"))
        if "缺租金" in gap_text or area.get("租金压力") in {"高", "中高", "待补"}:
            task_types.append(("租金招商核验", "租金、面积、转让费、物业费、扣点/保底、免租期"))
        if text(area.get("名称")) not in image_names:
            task_types.append(("实地图片核验", "商圈入口、主通道、门头连续面、餐饮楼层、竞品排队、停车外摆"))
        if not has_geo(area):
            task_types.append(("地图坐标核验", "补经纬度、行政区、街道、1/1.5/3公里半径"))
        if not task_types:
            task_types.append(("复核确认", "抽样核验来源、更新时间和评分依据"))
        for task_type, content in task_types[:3]:
            tasks.append({
                "任务ID": f"TASK-V04-{task_index:04d}",
                "对象类型": "商圈",
                "对象ID": area.get("记录ID"),
                "省份": area.get("省份"),
                "城市": area.get("城市"),
                "区县": area.get("区县"),
                "任务类型": task_type,
                "优先级": priority,
                "核验内容": content,
                "缺口字段": gap_text,
                "建议负责人": "拓展/选址/加盟支持",
                "状态": "待处理",
                "截止建议": (datetime.now() + timedelta(days=7 if priority == "P1" else 14)).strftime("%Y-%m-%d"),
                "关联商圈": area.get("名称"),
                "备注": f"当前质量分{quality}，完成后回写数据质量评分。",
                "数据更新时间": TODAY,
            })
            task_index += 1
    return tasks


def build_candidates(areas, districts, streets, radius_rows):
    radius_by_id = {row["对象ID"]: row for row in radius_rows}
    pool = []
    for row in areas:
        if row.get("省份") == "福建":
            pool.append(("商圈", row, row.get("名称")))
    for row in districts:
        if row.get("省份") == "福建":
            pool.append(("区县", row, row.get("区县")))
    for row in streets:
        if row.get("省份") == "福建" and row.get("城市") in {"福州", "厦门", "泉州"}:
            pool.append(("街道", row, row.get("街道/片区")))
    rows = []
    for index, (obj_type, row, name) in enumerate(sorted(pool, key=lambda item: score_value(item[1]), reverse=True)[:90], start=1):
        radius = radius_by_id.get(row.get("记录ID"))
        quality, gaps = classify_quality(row, radius)
        score = score_value(row)
        conclusion = decision(score, quality)
        fit = store_fit(row)
        for store_type in fit.split("、")[:2]:
            rows.append({
                "候选ID": f"CAND-V04-{index:04d}-{store_type}",
                "对象类型": obj_type,
                "对象ID": row.get("记录ID"),
                "省份": row.get("省份"),
                "城市": row.get("城市"),
                "区县": row.get("区县"),
                "名称": name,
                "店型模式": store_type,
                "推荐结论": conclusion,
                "推荐等级": row.get("推荐等级") or row.get("最高推荐等级"),
                "综合评分": score,
                "数据质量评分": quality,
                "主要依据": row.get("主要依据") or row.get("优先逻辑") or row.get("商业成熟度"),
                "主要风险": row.get("主要风险") or gaps,
                "下一步动作": "加入本周核验清单，补平台/租金/实地照片" if quality < 75 else "可进入点位搜索和物业接触",
                "收藏状态": "未收藏",
                "数据更新时间": TODAY,
            })
    return rows


def build_reports(candidates, radius_rows):
    radius_by_name = {row.get("名称"): row for row in radius_rows}
    reports = []
    for index, cand in enumerate([row for row in candidates if row["对象类型"] == "商圈"][:36], start=1):
        radius = radius_by_name.get(cand["名称"], {})
        reports.append({
            "报告ID": f"REPORT-V04-{index:04d}",
            "对象类型": cand["对象类型"],
            "对象ID": cand["对象ID"],
            "标题": f"{cand['城市']}{cand['名称']}选址核验摘要",
            "结论": f"{cand['推荐结论']}，适合{cand['店型模式']}，综合评分{cand['综合评分']}，质量分{cand['数据质量评分']}。",
            "地图摘要": f"经纬度与半径为V0.4近似计算；1.5公里POI {radius.get('1.5公里POI数', '待补')}，3公里竞品 {radius.get('3公里竞品数', '待补')}。",
            "评分摘要": cand["主要依据"],
            "证据摘要": f"{radius.get('堂食半径判断', '堂食半径待补')}；{radius.get('外卖半径判断', '外卖半径待补')}。",
            "风险摘要": cand["主要风险"],
            "下一步动作": cand["下一步动作"],
            "生成状态": "可预览",
            "数据更新时间": TODAY,
        })
    return reports


def update_existing_rows(wb, radius_rows):
    radius_by_id = {row["对象ID"]: row for row in radius_rows}
    target_sheets = ["城市库", "区县库", "商圈库", "街道库", "点位库", "POI库", "租金样本库", "竞品门店库"]
    for title in target_sheets:
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        ensure_headers(ws, QUALITY_COLUMNS)
        headers = header_map(ws)
        for row_idx in range(2, ws.max_row + 1):
            record = {header: ws.cell(row=row_idx, column=col).value for header, col in headers.items()}
            radius = radius_by_id.get(value(record.get("记录ID")))
            quality, gaps = classify_quality(record, radius)
            score = score_value(record)
            ws.cell(row=row_idx, column=headers["数据质量评分"]).value = quality
            ws.cell(row=row_idx, column=headers["数据缺口"]).value = gaps
            ws.cell(row=row_idx, column=headers["是否过期"]).value = "否"
            ws.cell(row=row_idx, column=headers["是否冲突"]).value = "否"
            ws.cell(row=row_idx, column=headers["核验动作"]).value = "补平台截图、地图API、租金报价、实地照片" if quality < 78 else "抽样复核"
            ws.cell(row=row_idx, column=headers["决策结论"]).value = decision(score, quality)
            ws.cell(row=row_idx, column=headers["店型匹配"]).value = store_fit(record)
            ws.cell(row=row_idx, column=headers["半径统计ID"]).value = radius["记录ID"] if radius else "待补"


def main():
    workbook = find_workbook()
    wb = load_workbook(workbook)

    radius_headers = ["记录ID", "对象类型", "对象ID", "省份", "城市", "区县", "名称", "经度", "纬度", "1公里POI数", "1公里竞品数", "1公里租金样本数", "1.5公里POI数", "1.5公里竞品数", "1.5公里租金样本数", "3公里POI数", "3公里竞品数", "3公里租金样本数", "堂食半径判断", "外卖半径判断", "最近竞品", "最近竞品距离km", "最近租金", "数据质量评分", "数据缺口", "核验动作", "数据更新时间", "来源", "验证状态"]
    task_headers = ["任务ID", "对象类型", "对象ID", "省份", "城市", "区县", "任务类型", "优先级", "核验内容", "缺口字段", "建议负责人", "状态", "截止建议", "关联商圈", "备注", "数据更新时间"]
    candidate_headers = ["候选ID", "对象类型", "对象ID", "省份", "城市", "区县", "名称", "店型模式", "推荐结论", "推荐等级", "综合评分", "数据质量评分", "主要依据", "主要风险", "下一步动作", "收藏状态", "数据更新时间"]
    report_headers = ["报告ID", "对象类型", "对象ID", "标题", "结论", "地图摘要", "评分摘要", "证据摘要", "风险摘要", "下一步动作", "生成状态", "数据更新时间"]
    weight_headers = ["店型模式", "对象类型", "维度", "权重", "说明"]

    radius_ws = ensure_sheet(wb, "半径统计表", radius_headers)
    task_ws = ensure_sheet(wb, "核验任务表", task_headers)
    candidate_ws = ensure_sheet(wb, "候选清单表", candidate_headers)
    report_ws = ensure_sheet(wb, "报告输出表", report_headers)
    weight_ws = ensure_sheet(wb, "店型权重表", weight_headers)

    for ws, prefixes in [
        (radius_ws, ("RAD-V04-",)),
        (task_ws, ("TASK-V04-",)),
        (candidate_ws, ("CAND-V04-",)),
        (report_ws, ("REPORT-V04-",)),
    ]:
        clear_generated(ws, prefixes)
    for row in range(weight_ws.max_row, 1, -1):
        weight_ws.delete_rows(row)

    areas = records(wb["商圈库"])
    districts = records(wb["区县库"]) if "区县库" in wb.sheetnames else []
    streets = records(wb["街道库"]) if "街道库" in wb.sheetnames else []
    poi_rows = records(wb["POI库"]) if "POI库" in wb.sheetnames else []
    competitor_rows = records(wb["竞品门店库"]) if "竞品门店库" in wb.sheetnames else []
    rent_rows = records(wb["租金样本库"]) if "租金样本库" in wb.sheetnames else []
    image_rows = records(wb["图片素材库"]) if "图片素材库" in wb.sheetnames else []

    radius_rows = build_radius_rows(areas, poi_rows, competitor_rows, rent_rows)
    task_rows = build_tasks(areas, radius_rows, image_rows)
    candidate_rows = build_candidates(areas, districts, streets, radius_rows)
    report_rows = build_reports(candidate_rows, radius_rows)

    append_records(radius_ws, radius_rows)
    append_records(task_ws, task_rows)
    append_records(candidate_ws, candidate_rows)
    append_records(report_ws, report_rows)
    append_records(weight_ws, [
        {"店型模式": row[0], "对象类型": row[1], "维度": row[2], "权重": row[3], "说明": row[4]}
        for row in STORE_TYPE_WEIGHTS
    ])
    update_existing_rows(wb, radius_rows)

    wb.save(workbook)
    print(f"已完成V0.4数据升级：{workbook}")
    print(f"半径统计 {len(radius_rows)}，核验任务 {len(task_rows)}，候选清单 {len(candidate_rows)}，报告摘要 {len(report_rows)}")


if __name__ == "__main__":
    main()
