from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
REQUIRED_SHEETS = {"城市库", "商圈库", "点位库", "数据源登记表"}
TODAY = datetime(2026, 5, 10)
TODAY_TEXT = TODAY.strftime("%Y-%m-%d")

MISSING_TOKENS = (
    "待补",
    "待授权",
    "待实地",
    "待招商",
    "待平台",
    "待高德",
    "待百度",
    "待走访",
    "暂无",
    "无法计算",
)

SOURCE_LEVELS = [
    ("L5", "实地验证", 1.0, "实地照片、现场走访、房东/招商报价、合同条件", "可直接作为选址判断证据，但仍需定期复核。"),
    ("L4", "官方数据", 0.9, "统计公报、统计年鉴、政府公开数据", "适合城市/区县基本盘判断，不能替代点位现场判断。"),
    ("L3", "平台截图/API", 0.8, "美团/点评截图、高德/百度API、授权平台导出", "适合竞品、POI、热力、半径与榜单核验。"),
    ("L2", "公开资料", 0.6, "招商资料、新闻、商场公开信息、公开地图资料", "可作为线索，关键结论需平台或实地复核。"),
    ("L1", "AI线索", 0.4, "AI整理、经验推断、未核验结构化线索", "只能用于发现方向，不能直接推荐点位。"),
    ("L0", "缺失", 0.0, "空字段、待补、待授权、待实地观察、待招商报价", "必须进入核验任务，禁止直接推荐。"),
]

SOURCE_WEIGHT = {code: weight for code, _, weight, _, _ in SOURCE_LEVELS}

OBJECT_REQUIREMENTS = {
    "城市库": ["推荐等级", "综合评分", "经度", "纬度", "常住人口", "GDP/社零/第三产业", "餐饮消费活跃度", "核心商圈", "主要依据", "主要风险", "下一步动作", "来源类型"],
    "区县库": ["推荐等级", "综合评分", "经度", "纬度", "常住人口/人口密度", "商业成熟度", "餐饮密度", "竞品分布", "消费力", "主要依据", "主要风险", "下一步动作", "来源类型"],
    "商圈库": ["推荐等级", "综合评分", "经度", "纬度", "商圈类型", "覆盖街道", "人流热力", "消费场景", "餐饮集中度", "竞品表现", "租金压力", "3公里配送覆盖", "适合店型", "主要依据", "主要风险", "下一步动作", "来源类型"],
    "街道库": ["最高推荐等级", "环境评分", "经度", "纬度", "关联商圈", "人流/可见性", "餐饮集聚", "住宅", "办公", "停车", "竞品", "租金", "堂食1-1.5公里判断", "外卖3公里判断", "主要依据", "主要风险", "下一步动作"],
    "POI库": ["经度", "纬度", "POI类型", "关联商圈", "客流/覆盖判断", "堂食半径价值", "外卖半径价值", "选址意义", "来源类型", "验证状态"],
    "竞品门店库": ["经度", "纬度", "竞品品牌", "关联商圈", "平台", "评分", "评论数", "人均", "月销量/热度", "外卖表现", "堂食表现", "竞争压力", "来源类型", "验证状态"],
    "租金样本库": ["经度", "纬度", "关联商圈", "铺位类型", "租金区间", "面积建议", "转让费", "物业费", "合同条件", "回本压力", "来源类型", "验证状态"],
    "图片素材库": ["关联对象", "图片类型", "图片路径/链接", "拍摄日期", "来源", "说明"],
    "候选清单表": ["对象类型", "对象ID", "经度", "纬度", "店型模式", "推荐结论", "推荐等级", "综合评分", "数据质量评分", "主要依据", "主要风险", "下一步动作"],
}

V05_EXISTING_HEADERS = [
    "来源等级",
    "字段完整率",
    "可信度标签",
    "严格决策结论",
    "推荐门槛说明",
    "V0.5核验状态",
    "证据摘要",
]


def find_workbook() -> Path:
    candidates = []
    for path in Path("C:/CodexData").rglob("*.xlsx"):
        if ".bak-" in path.name:
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if REQUIRED_SHEETS.issubset(set(wb.sheetnames)):
                bonus = sum(1 for sheet in ["数据源等级表", "字段核验矩阵", "城市KPI表", "铺位线索表"] if sheet in wb.sheetnames)
                candidates.append((bonus, path.stat().st_mtime, path))
        except Exception:
            continue
    if candidates:
        return sorted(candidates, reverse=True)[0][2]
    raise FileNotFoundError("未找到周麻婆选址数据总表")


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def display(value) -> str:
    value = text(value)
    return value if value else "待补"


def is_missing(value) -> bool:
    value = text(value)
    if not value:
        return True
    if value.lower() in {"none", "nan", "null"}:
        return True
    return any(token in value for token in MISSING_TOKENS)


def to_float(value, default=0.0) -> float:
    value = text(value)
    if not value:
        return default
    keep = "".join(ch for ch in value if ch.isdigit() or ch == ".")
    try:
        return float(keep) if keep else default
    except Exception:
        return default


def sheet_records(ws):
    headers = [text(cell.value) for cell in ws[1]]
    records = []
    for row_idx in range(2, ws.max_row + 1):
        record = {header: ws.cell(row_idx, col_idx).value for col_idx, header in enumerate(headers, 1) if header}
        if any(not is_missing(value) for value in record.values()):
            records.append((row_idx, record))
    return headers, records


def ensure_headers(ws, headers):
    current = [text(cell.value) for cell in ws[1]]
    last_col = len(current)
    for header in headers:
        if header not in current:
            last_col += 1
            ws.cell(1, last_col).value = header
            current.append(header)
    return current


def row_value(row, keys, default=""):
    for key in keys:
        if key in row and not is_missing(row.get(key)):
            return row.get(key)
    return default


def source_level(row) -> str:
    path = row_value(row, ["图片路径/链接", "资料路径", "截图路径", "文件路径"])
    blob = " ".join(text(row.get(key)) for key in ["来源", "来源类型", "验证状态", "资料路径", "图片路径/链接", "平台", "备注", "docx原始数据"])
    if not blob and is_missing(path):
        return "L0"
    if not is_missing(path) and any(ext in text(path).lower() for ext in [".jpg", ".jpeg", ".png", ".webp", ".pdf"]):
        return "L5"
    if "实地" in blob and "待" not in blob:
        return "L5"
    if any(token in blob for token in ["官方", "统计", "公报", "年鉴", "政府", ".gov"]):
        return "L4"
    if any(token in blob for token in ["平台截图", "截图", "美团", "点评", "高德", "百度", "地图API", "授权", "招商报价"]):
        return "L3"
    if any(token in blob for token in ["公开", "DataV", "docx", "招商资料", "新闻", "商场公开"]):
        return "L2"
    if any(token in blob for token in ["AI", "线索", "推断"]):
        return "L1"
    return "L1"


def quality_label(value: float) -> str:
    if value >= 85:
        return "高"
    if value >= 70:
        return "中高"
    if value >= 55:
        return "中"
    if value >= 40:
        return "低"
    return "待补"


def strict_decision(score: float, quality: float) -> tuple[str, str]:
    if quality < 50:
        return "先补数据", "质量分低于50：只进入核验，不进入推荐池。"
    if quality < 70:
        if score >= 78:
            return "潜力推荐/待核验", "质量分低于70：禁止直接推荐，只能作为潜力对象安排核验。"
        return "谨慎/待核验", "质量分低于70：需补平台、租金、照片或实地证据。"
    if score >= 85:
        return "推荐", "质量分达到70且评分较高，可进入推荐池。"
    if score >= 75:
        return "谨慎推荐", "质量分达到70，但仍需看租金、竞品或物业条件。"
    if score >= 60:
        return "谨慎", "评分未达到推荐门槛，适合保留观察。"
    return "先补数据", "评分和质量均不足，先补基础数据。"


def score_grade(score: float) -> str:
    if score >= 85:
        return "A"
    if score >= 75:
        return "B"
    if score >= 65:
        return "C"
    if score > 0:
        return "D"
    return "待定"


def row_score(row) -> float:
    return to_float(row_value(row, ["综合评分", "环境评分", "评分"]), 0)


def row_identity(row):
    return display(row_value(row, ["记录ID", "候选ID", "任务ID", "报告ID", "线索ID", "对象ID", "名称"]))


def row_name(row):
    return display(row_value(row, ["名称", "区县", "街道/片区", "关联商圈", "铺位线索名称", "门店名称", "城市"]))


def related_rows(rows, city, name, district=""):
    city = text(city)
    name = text(name)
    district = text(district)
    result = []
    for row in rows:
        if city and text(row.get("城市")) != city:
            continue
        haystack = " ".join(text(row.get(key)) for key in ["名称", "关联商圈", "关联对象", "街道/片区", "区县"])
        if name and name in haystack:
            result.append(row)
        elif district and text(row.get("区县")) == district:
            result.append(row)
    return result


def related_exists(rows, city, name, district="", require_real=True):
    for row in related_rows(rows, city, name, district):
        if require_real:
            if not is_missing(row_value(row, ["评分", "评论数", "人均", "月销量/热度", "租金区间", "图片路径/链接", "经度"])):
                return True
        else:
            return True
    return False


def field_coverage(row, requirements):
    if not requirements:
        return 0
    present = sum(0 if is_missing(row.get(field)) else 1 for field in requirements)
    return round(present / len(requirements) * 100, 1)


def evidence_coverage(sheet_name, row, indexes):
    city = text(row.get("城市"))
    district = text(row.get("区县"))
    name = row_name(row)
    checks = []
    if sheet_name in {"商圈库", "街道库", "区县库", "候选清单表"}:
        checks.extend(
            [
                not is_missing(row.get("经度")) and not is_missing(row.get("纬度")),
                related_exists(indexes["competitors"], city, name, district, require_real=True),
                related_exists(indexes["rent"], city, name, district, require_real=True),
                related_exists(indexes["images"], city, name, district, require_real=True),
                related_exists(indexes["radius"], city, name, district, require_real=False),
            ]
        )
    elif sheet_name == "竞品门店库":
        checks.extend([not is_missing(row.get("评分")), not is_missing(row.get("评论数")), not is_missing(row.get("月销量/热度")), not is_missing(row.get("经度"))])
    elif sheet_name == "租金样本库":
        checks.extend([not is_missing(row.get("租金区间")), not is_missing(row.get("转让费")), not is_missing(row.get("物业费")), not is_missing(row.get("合同条件")), not is_missing(row.get("经度"))])
    elif sheet_name == "图片素材库":
        checks.extend([not is_missing(row.get("图片路径/链接")), not is_missing(row.get("拍摄日期")), not is_missing(row.get("关联对象"))])
    else:
        checks.extend([not is_missing(row.get("经度")), not is_missing(row.get("纬度")), not is_missing(row.get("来源类型"))])
    return round(sum(1 for flag in checks if flag) / max(1, len(checks)) * 100, 1)


def quality_score(sheet_name, row, indexes):
    level = source_level(row)
    source_part = SOURCE_WEIGHT[level] * 25
    requirements = OBJECT_REQUIREMENTS.get(sheet_name, [])
    coverage = field_coverage(row, requirements)
    evidence = evidence_coverage(sheet_name, row, indexes)
    quality = source_part + coverage * 0.45 + evidence * 0.30
    if level == "L1":
        quality = min(quality, 68)
    if level == "L0":
        quality = min(quality, 45)
    if sheet_name in {"商圈库", "街道库", "候选清单表"}:
        city = text(row.get("城市"))
        district = text(row.get("区县"))
        name = row_name(row)
        no_platform = not related_exists(indexes["competitors"], city, name, district, require_real=True)
        no_rent = not related_exists(indexes["rent"], city, name, district, require_real=True)
        no_photo = not related_exists(indexes["images"], city, name, district, require_real=True)
        if no_platform and no_rent and no_photo:
            quality = min(quality, 66)
        elif no_photo:
            quality = min(quality, 76)
    return round(max(0, min(100, quality)), 1), coverage, evidence, level


def row_gaps(sheet_name, row, indexes):
    requirements = OBJECT_REQUIREMENTS.get(sheet_name, [])
    gaps = [field for field in requirements if is_missing(row.get(field))]
    city = text(row.get("城市"))
    district = text(row.get("区县"))
    name = row_name(row)
    if sheet_name in {"商圈库", "街道库", "候选清单表"}:
        if not related_exists(indexes["competitors"], city, name, district, require_real=True):
            gaps.append("缺平台竞品截图/销量")
        if not related_exists(indexes["rent"], city, name, district, require_real=True):
            gaps.append("缺租金报价/合同条件")
        if not related_exists(indexes["images"], city, name, district, require_real=True):
            gaps.append("缺实地照片")
        if not related_exists(indexes["radius"], city, name, district, require_real=False):
            gaps.append("缺半径统计")
    return "；".join(dict.fromkeys(gaps)) if gaps else "核心字段已覆盖，仍需定期复核"


def action_for_gap(gap: str) -> str:
    if "竞品" in gap or "平台" in gap or "销量" in gap:
        return "补美团/点评授权截图，记录评分、评论、人均、销量、榜单和距离"
    if "租金" in gap or "合同" in gap or "物业" in gap or "转让" in gap:
        return "联系招商/房东/中介，补租金、面积、转让费、物业费和合同条件"
    if "照片" in gap or "图片" in gap:
        return "按五类照片标准走访拍摄：入口、门头、竞品、业态、停车交通"
    if "经度" in gap or "纬度" in gap or "半径" in gap:
        return "用高德/百度坐标复核并重算1km、1.5km、3km半径"
    return "人工复核字段并回写来源、日期、置信度"


def set_cell(ws, headers, row_idx, header, value):
    if header not in headers:
        headers = ensure_headers(ws, [header])
    col = headers.index(header) + 1
    ws.cell(row_idx, col).value = value


def build_indexes(all_records):
    return {
        "competitors": [row for _, row in all_records.get("竞品门店库", [])],
        "rent": [row for _, row in all_records.get("租金样本库", [])],
        "images": [row for _, row in all_records.get("图片素材库", [])],
        "radius": [row for _, row in all_records.get("半径统计表", [])],
    }


def record_lookup(all_records):
    lookup = {}
    for sheet, records in all_records.items():
        for _, row in records:
            for key in [row_value(row, ["记录ID", "对象ID", "候选ID"]), row_name(row)]:
                key = text(key)
                if key:
                    lookup[key] = row
    return lookup


def strict_for_row(row):
    score = row_score(row)
    quality = to_float(row.get("数据质量评分"), 0)
    if not score and quality:
        score = quality
    return strict_decision(score, quality)


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def rewrite_sheet(wb, title, headers, rows):
    if title in wb.sheetnames:
        idx = wb.sheetnames.index(title)
        del wb[title]
        ws = wb.create_sheet(title, idx)
    else:
        ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    return ws


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="EAF3EF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="173328")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = col[0].column_letter
        header = text(ws[f"{letter}1"].value)
        if header in {"主要依据", "主要风险", "下一步动作", "数据缺口", "核验内容", "核验动作", "证据摘要", "备注"}:
            ws.column_dimensions[letter].width = 34
        elif len(header) >= 6:
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 13


def refresh_records(wb):
    all_records = {}
    headers_by_sheet = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers, records = sheet_records(ws)
        headers_by_sheet[sheet] = headers
        all_records[sheet] = records
    return headers_by_sheet, all_records


def upgrade_existing_sheets(wb):
    headers_by_sheet, all_records = refresh_records(wb)
    indexes = build_indexes(all_records)
    core_sheets = ["城市库", "区县库", "商圈库", "街道库", "POI库", "竞品门店库", "租金样本库", "图片素材库"]
    for sheet_name in core_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = ensure_headers(ws, V05_EXISTING_HEADERS)
        _, records = sheet_records(ws)
        for row_idx, row in records:
            quality, coverage, evidence, level = quality_score(sheet_name, row, indexes)
            score = row_score(row)
            if not score and quality:
                score = quality
            decision, threshold_note = strict_decision(score, quality)
            gaps = row_gaps(sheet_name, row, indexes)
            set_cell(ws, headers, row_idx, "来源等级", level)
            set_cell(ws, headers, row_idx, "字段完整率", coverage)
            set_cell(ws, headers, row_idx, "数据质量评分", quality)
            set_cell(ws, headers, row_idx, "可信度标签", quality_label(quality))
            set_cell(ws, headers, row_idx, "严格决策结论", decision)
            set_cell(ws, headers, row_idx, "推荐门槛说明", threshold_note)
            set_cell(ws, headers, row_idx, "V0.5核验状态", "需补证据" if quality < 70 else "可复核使用")
            set_cell(ws, headers, row_idx, "数据缺口", gaps)
            set_cell(ws, headers, row_idx, "核验动作", action_for_gap(gaps))
            set_cell(ws, headers, row_idx, "决策结论", decision)
            set_cell(ws, headers, row_idx, "证据摘要", f"来源{level}，字段完整率{coverage}%，证据覆盖率{evidence}%。")
        style_sheet(ws)
    return refresh_records(wb)


def upgrade_candidates(wb, all_records):
    if "候选清单表" not in wb.sheetnames:
        return
    lookup = record_lookup(all_records)
    ws = wb["候选清单表"]
    headers = ensure_headers(ws, ["经度", "纬度", *V05_EXISTING_HEADERS, "收藏分组"])
    _, records = sheet_records(ws)
    for row_idx, row in records:
        target = lookup.get(text(row.get("对象ID"))) or lookup.get(row_name(row))
        if target:
            if is_missing(row.get("经度")) and not is_missing(target.get("经度")):
                set_cell(ws, headers, row_idx, "经度", target.get("经度"))
            if is_missing(row.get("纬度")) and not is_missing(target.get("纬度")):
                set_cell(ws, headers, row_idx, "纬度", target.get("纬度"))
            row = {**target, **row, "经度": target.get("经度", row.get("经度")), "纬度": target.get("纬度", row.get("纬度"))}
        quality = to_float(row.get("数据质量评分"), to_float(target.get("数据质量评分") if target else "", 0))
        score = row_score(row)
        decision, threshold_note = strict_decision(score, quality)
        level = source_level(target or row)
        gaps = row.get("数据缺口") or (target or {}).get("数据缺口") or "缺平台截图、租金报价、真实照片和地图API复核"
        set_cell(ws, headers, row_idx, "来源等级", level)
        set_cell(ws, headers, row_idx, "字段完整率", field_coverage(row, OBJECT_REQUIREMENTS["候选清单表"]))
        set_cell(ws, headers, row_idx, "可信度标签", quality_label(quality))
        set_cell(ws, headers, row_idx, "严格决策结论", decision)
        set_cell(ws, headers, row_idx, "推荐结论", decision)
        set_cell(ws, headers, row_idx, "推荐门槛说明", threshold_note)
        set_cell(ws, headers, row_idx, "V0.5核验状态", "需补证据" if quality < 70 else "可复核使用")
        set_cell(ws, headers, row_idx, "证据摘要", f"继承对象坐标与质量；来源{level}。")
        set_cell(ws, headers, row_idx, "数据缺口", gaps)
        set_cell(ws, headers, row_idx, "收藏分组", row.get("收藏分组") or "关注")
    style_sheet(ws)


def build_field_matrix(all_records):
    rows = []
    n = 1
    for sheet_name, requirements in OBJECT_REQUIREMENTS.items():
        for _, row in all_records.get(sheet_name, []):
            level = source_level(row)
            quality = to_float(row.get("数据质量评分"), 0)
            for field in requirements:
                missing = is_missing(row.get(field))
                priority = "P1" if missing and quality >= 70 else ("P2" if missing else "已覆盖")
                rows.append(
                    {
                        "矩阵ID": f"MATRIX-V05-{n:05d}",
                        "对象类型": sheet_name.replace("库", "").replace("表", ""),
                        "对象ID": row_identity(row),
                        "省份": display(row.get("省份")),
                        "城市": display(row.get("城市")),
                        "区县": display(row.get("区县")),
                        "名称": row_name(row),
                        "字段": field,
                        "当前值": display(row.get(field)),
                        "是否缺失": "是" if missing else "否",
                        "来源等级": level,
                        "字段权重": "关键" if field in {"经度", "纬度", "租金区间", "评分", "评论数", "图片路径/链接", "主要风险"} else "普通",
                        "核验优先级": priority,
                        "核验动作": action_for_gap(field) if missing else "定期复核",
                        "数据更新时间": TODAY_TEXT,
                    }
                )
                n += 1
    return rows


def field_gap_counts(matrix_rows, city="", district="", name=""):
    result = Counter()
    for row in matrix_rows:
        if row["是否缺失"] != "是":
            continue
        if city and row["城市"] != city:
            continue
        if district and row["区县"] != district:
            continue
        if name and row["名称"] != name:
            continue
        field = row["字段"]
        if any(token in field for token in ["竞品", "评分", "评论", "销量"]):
            result["缺竞品"] += 1
        if any(token in field for token in ["租金", "合同", "物业", "转让"]):
            result["缺租金"] += 1
        if any(token in field for token in ["图片", "照片", "拍摄"]):
            result["缺照片"] += 1
        if any(token in field for token in ["平台", "评分", "评论", "销量"]):
            result["缺平台数据"] += 1
        if any(token in field for token in ["经度", "纬度", "半径"]):
            result["缺坐标/半径"] += 1
    return result


def build_kpis(all_records, matrix_rows):
    cities = [row for _, row in all_records.get("城市库", [])]
    districts = [row for _, row in all_records.get("区县库", [])]
    areas = [row for _, row in all_records.get("商圈库", [])]
    streets = [row for _, row in all_records.get("街道库", [])]
    poi = [row for _, row in all_records.get("POI库", [])]
    competitors = [row for _, row in all_records.get("竞品门店库", [])]
    rent = [row for _, row in all_records.get("租金样本库", [])]
    tasks = [row for _, row in all_records.get("核验任务表", [])]
    candidates = [row for _, row in all_records.get("候选清单表", [])]

    city_rows = []
    for city in cities:
        city_name = text(city.get("城市"))
        city_area = [row for row in areas if text(row.get("城市")) == city_name]
        city_district = [row for row in districts if text(row.get("城市")) == city_name]
        city_candidates = [row for row in candidates if text(row.get("城市")) == city_name]
        city_tasks = [row for row in tasks if text(row.get("城市")) == city_name and text(row.get("状态")) != "已完成"]
        qualities = [to_float(row.get("数据质量评分"), 0) for row in city_area + city_district if to_float(row.get("数据质量评分"), 0)]
        gaps = field_gap_counts(matrix_rows, city=city_name)
        city_rows.append(
            {
                "省份": display(city.get("省份")),
                "城市": city_name,
                "推荐等级": display(city.get("推荐等级")),
                "综合评分": display(city.get("综合评分")),
                "严格决策结论": display(city.get("严格决策结论") or city.get("决策结论")),
                "来源等级": display(city.get("来源等级")),
                "平均数据质量分": round(sum(qualities) / max(1, len(qualities)), 1),
                "区县数": len(city_district),
                "商圈数": len(city_area),
                "街道数": sum(1 for row in streets if text(row.get("城市")) == city_name),
                "POI数": sum(1 for row in poi if text(row.get("城市")) == city_name),
                "竞品数": sum(1 for row in competitors if text(row.get("城市")) == city_name),
                "租金样本数": sum(1 for row in rent if text(row.get("城市")) == city_name),
                "核验任务数": len(city_tasks),
                "P1任务数": sum(1 for row in city_tasks if text(row.get("优先级")) == "P1"),
                "推荐对象数": sum(1 for row in city_candidates if text(row.get("推荐结论")) == "推荐"),
                "潜力推荐数": sum(1 for row in city_candidates if "潜力推荐" in text(row.get("推荐结论"))),
                "缺竞品": gaps["缺竞品"],
                "缺租金": gaps["缺租金"],
                "缺照片": gaps["缺照片"],
                "缺平台数据": gaps["缺平台数据"],
                "数据缺口摘要": "；".join(k for k, v in gaps.items() if v) or "核心字段已覆盖，仍需复核",
                "下一步动作": "福州优先完成20个商圈的竞品、租金、照片核验" if city_name == "福州" else "保留骨架，按重点城市节奏补商圈和街道数据",
                "数据更新时间": TODAY_TEXT,
            }
        )

    district_rows = []
    for district in districts:
        city_name = text(district.get("城市"))
        district_name = text(district.get("区县"))
        area_rows = [row for row in areas if text(row.get("城市")) == city_name and text(row.get("区县")) == district_name]
        district_tasks = [row for row in tasks if text(row.get("城市")) == city_name and text(row.get("区县")) == district_name and text(row.get("状态")) != "已完成"]
        qualities = [to_float(row.get("数据质量评分"), 0) for row in area_rows if to_float(row.get("数据质量评分"), 0)]
        gaps = field_gap_counts(matrix_rows, city=city_name, district=district_name)
        district_rows.append(
            {
                "省份": display(district.get("省份")),
                "城市": city_name,
                "区县": district_name,
                "推荐等级": display(district.get("推荐等级")),
                "综合评分": display(district.get("综合评分")),
                "严格决策结论": display(district.get("严格决策结论") or district.get("决策结论")),
                "来源等级": display(district.get("来源等级")),
                "平均数据质量分": round(sum(qualities) / max(1, len(qualities)), 1) if qualities else display(district.get("数据质量评分")),
                "商圈数": len(area_rows),
                "街道数": sum(1 for row in streets if text(row.get("城市")) == city_name and text(row.get("区县")) == district_name),
                "POI数": sum(1 for row in poi if text(row.get("城市")) == city_name and text(row.get("区县")) == district_name),
                "竞品数": sum(1 for row in competitors if text(row.get("城市")) == city_name and text(row.get("区县")) == district_name),
                "租金样本数": sum(1 for row in rent if text(row.get("城市")) == city_name and text(row.get("区县")) == district_name),
                "核验任务数": len(district_tasks),
                "P1任务数": sum(1 for row in district_tasks if text(row.get("优先级")) == "P1"),
                "缺竞品": gaps["缺竞品"],
                "缺租金": gaps["缺租金"],
                "缺照片": gaps["缺照片"],
                "缺平台数据": gaps["缺平台数据"],
                "数据缺口摘要": "；".join(k for k, v in gaps.items() if v) or "核心字段已覆盖，仍需复核",
                "下一步动作": "筛选3-5个高分商圈进入本周走访" if area_rows else "补区县商圈骨架",
                "数据更新时间": TODAY_TEXT,
            }
        )
    return city_rows, district_rows


def build_business_profiles(all_records, matrix_rows):
    areas = [row for _, row in all_records.get("商圈库", [])]
    radius = [row for _, row in all_records.get("半径统计表", [])]
    competitors = [row for _, row in all_records.get("竞品门店库", [])]
    rent = [row for _, row in all_records.get("租金样本库", [])]
    images = [row for _, row in all_records.get("图片素材库", [])]
    radius_by_id = {text(row.get("对象ID")): row for row in radius}
    rows = []
    for area in areas:
        name = row_name(area)
        city = text(area.get("城市"))
        district = text(area.get("区县"))
        related_comp = related_rows(competitors, city, name, district)
        related_rent = related_rows(rent, city, name, district)
        related_img = related_rows(images, city, name, district)
        rad = radius_by_id.get(text(area.get("记录ID"))) or next(iter(related_rows(radius, city, name, district)), {})
        gaps = field_gap_counts(matrix_rows, city=city, district=district, name=name)
        score = row_score(area)
        quality = to_float(area.get("数据质量评分"), 0)
        decision, threshold_note = strict_decision(score, quality)
        rows.append(
            {
                "画像ID": f"PROFILE-{row_identity(area)}",
                "省份": display(area.get("省份")),
                "城市": city,
                "区县": display(area.get("区县")),
                "街道": display(area.get("街道") or area.get("覆盖街道")),
                "商圈名称": name,
                "商圈类型": display(area.get("商圈类型")),
                "推荐等级": display(area.get("推荐等级")),
                "综合评分": display(area.get("综合评分")),
                "严格决策结论": decision,
                "数据质量评分": quality,
                "来源等级": display(area.get("来源等级")),
                "质量门槛说明": threshold_note,
                "1.5公里POI数": display(rad.get("1.5公里POI数")),
                "3公里POI数": display(rad.get("3公里POI数")),
                "3公里竞品数": display(rad.get("3公里竞品数")),
                "3公里租金样本数": display(rad.get("3公里租金样本数")),
                "竞品核验数": len(related_comp),
                "租金样本数": len(related_rent),
                "照片证据数": sum(1 for row in related_img if not is_missing(row.get("图片路径/链接"))),
                "缺竞品": "是" if gaps["缺竞品"] else "否",
                "缺租金": "是" if gaps["缺租金"] else "否",
                "缺照片": "是" if gaps["缺照片"] else "否",
                "缺平台数据": "是" if gaps["缺平台数据"] else "否",
                "主要依据": display(area.get("主要依据")),
                "主要风险": display(area.get("主要风险")),
                "下一步动作": display(area.get("下一步动作") or action_for_gap(display(area.get("数据缺口")))),
                "数据缺口": display(area.get("数据缺口")),
                "数据更新时间": TODAY_TEXT,
            }
        )
    return rows


def build_street_environment(all_records):
    streets = [row for _, row in all_records.get("街道库", [])]
    rows = []
    for street in streets:
        score = row_score(street)
        quality = to_float(street.get("数据质量评分"), 0)
        decision, threshold_note = strict_decision(score, quality)
        rows.append(
            {
                "环境ID": f"ENV-{row_identity(street)}",
                "省份": display(street.get("省份")),
                "城市": display(street.get("城市")),
                "区县": display(street.get("区县")),
                "街道/片区": row_name(street),
                "关联商圈": display(street.get("关联商圈")),
                "推荐等级": display(street.get("最高推荐等级") or street.get("推荐等级") or score_grade(score)),
                "环境评分": score,
                "严格决策结论": decision,
                "数据质量评分": quality,
                "来源等级": display(street.get("来源等级")),
                "质量门槛说明": threshold_note,
                "人流/可见性": display(street.get("人流/可见性")),
                "餐饮集聚": display(street.get("餐饮集聚")),
                "住宅": display(street.get("住宅")),
                "办公": display(street.get("办公")),
                "停车": display(street.get("停车")),
                "竞品": display(street.get("竞品")),
                "租金": display(street.get("租金")),
                "堂食判断": display(street.get("堂食1-1.5公里判断")),
                "外卖判断": display(street.get("外卖3公里判断")),
                "数据缺口": display(street.get("数据缺口")),
                "下一步动作": display(street.get("下一步动作") or street.get("核验动作")),
                "数据更新时间": TODAY_TEXT,
            }
        )
    return rows


def lead_type(index):
    if index % 3 == 0:
        return ("购物中心铺位线索", "购物中心店", "60-120㎡", "餐饮楼层/主通道/入口动线")
    if index % 3 == 1:
        return ("临街铺位线索", "商圈店", "70-130㎡", "一层临街/可见门头/可外摆优先")
    return ("社区底商线索", "社区店", "55-110㎡", "社区入口/住宅密度/晚餐复购优先")


def build_site_leads(all_records):
    areas = [row for _, row in all_records.get("商圈库", [])]
    key_areas = sorted(
        [row for row in areas if text(row.get("城市")) in {"福州", "厦门", "泉州"} and not is_missing(row.get("经度"))],
        key=lambda row: (text(row.get("城市")) != "福州", -row_score(row), row_name(row)),
    )
    rows = []
    counter = 1
    per_city_limit = {"福州": 20, "厦门": 10, "泉州": 10}
    city_counts = Counter()
    for area in key_areas:
        city = text(area.get("城市"))
        if city_counts[city] >= per_city_limit.get(city, 0):
            continue
        city_counts[city] += 1
        repeat = 2 if city == "福州" and city_counts[city] <= 20 else 1
        for sub_index in range(repeat):
            line_type, store_type, area_size, frontage = lead_type(counter + sub_index)
            score = max(55, row_score(area) - (4 if line_type == "购物中心铺位线索" else 2))
            quality = min(68, to_float(area.get("数据质量评分"), 0))
            decision, threshold_note = strict_decision(score, quality)
            rows.append(
                {
                    "线索ID": f"LEAD-V05-{counter:04d}",
                    "来源对象ID": display(area.get("记录ID")),
                    "省份": display(area.get("省份")),
                    "城市": city,
                    "区县": display(area.get("区县")),
                    "街道/片区": display(area.get("街道") or area.get("覆盖街道")),
                    "关联商圈": row_name(area),
                    "铺位线索名称": f"{row_name(area)}-{line_type}",
                    "线索类型": line_type,
                    "店型模式": store_type,
                    "经度": display(area.get("经度")),
                    "纬度": display(area.get("纬度")),
                    "面积建议": area_size,
                    "租金/招商状态": "待招商/房东报价",
                    "楼层/门头": frontage,
                    "推荐结论": decision,
                    "推荐等级": score_grade(score),
                    "综合评分": round(score, 1),
                    "数据质量评分": quality,
                    "来源等级": "L1",
                    "质量门槛说明": threshold_note,
                    "数据缺口": "缺具体门牌、面积、租金、转让费、合同条件、门头照片、竞品100米核验",
                    "核验任务": "现场找铺/招商报价/门头照片/100米竞品观察",
                    "负责人": "拓展经理待分配",
                    "状态": "待核验",
                    "下一步动作": "加入本周核验清单，走访后回写铺位级评分",
                    "图片要求": "商圈入口、候选门头左右邻居、最强竞品、街道业态、停车交通",
                    "数据更新时间": TODAY_TEXT,
                }
            )
            counter += 1
    return rows


def enhance_tasks(wb, all_records, site_leads):
    if "核验任务表" not in wb.sheetnames:
        return
    ws = wb["核验任务表"]
    headers = ensure_headers(ws, ["来源等级", "数据质量评分", "质量门槛说明", "任务状态分组", "关联铺位线索"])
    old_rows = []
    for _, row in sheet_records(ws)[1]:
        if not text(row.get("任务ID")).startswith("TASK-V05-"):
            old_rows.append(row)
    top_areas = sorted(
        [row for _, row in all_records.get("商圈库", []) if text(row.get("城市")) == "福州"],
        key=lambda row: (-row_score(row), row_name(row)),
    )[:20]
    generated = []
    task_no = 1
    for area in top_areas:
        quality = to_float(area.get("数据质量评分"), 0)
        priority = "P1" if row_score(area) >= 85 else "P2"
        for task_type, gap in [
            ("竞品平台核验", "评分/评论/人均/月销量/榜单/距离"),
            ("租金招商核验", "租金区间/面积/转让费/物业费/合同条件"),
            ("实地图片核验", "入口/门头/竞品/业态/停车交通照片"),
        ]:
            generated.append(
                {
                    "任务ID": f"TASK-V05-{task_no:04d}",
                    "对象类型": "商圈",
                    "对象ID": display(area.get("记录ID")),
                    "省份": display(area.get("省份")),
                    "城市": display(area.get("城市")),
                    "区县": display(area.get("区县")),
                    "任务类型": task_type,
                    "优先级": priority,
                    "核验内容": action_for_gap(gap),
                    "缺口字段": gap,
                    "建议负责人": "拓展经理/选址分析师",
                    "状态": "待处理",
                    "截止建议": (TODAY + timedelta(days=7 if priority == "P1" else 14)).strftime("%Y-%m-%d"),
                    "关联商圈": row_name(area),
                    "备注": f"V0.5质量门槛任务：当前质量分{quality}，完成后重新计算推荐结论。",
                    "数据更新时间": TODAY_TEXT,
                    "来源等级": display(area.get("来源等级")),
                    "数据质量评分": quality,
                    "质量门槛说明": display(area.get("推荐门槛说明")),
                    "任务状态分组": "本周重点" if priority == "P1" else "两周内补齐",
                    "关联铺位线索": "待从铺位线索表选择",
                }
            )
            task_no += 1
    for lead in site_leads[:40]:
        generated.append(
            {
                "任务ID": f"TASK-V05-{task_no:04d}",
                "对象类型": "铺位线索",
                "对象ID": display(lead.get("线索ID")),
                "省份": display(lead.get("省份")),
                "城市": display(lead.get("城市")),
                "区县": display(lead.get("区县")),
                "任务类型": "铺位级现场核验",
                "优先级": "P1" if text(lead.get("城市")) == "福州" else "P2",
                "核验内容": "确认门牌、面积、租金、门头、动线、合同条件和100米竞品。",
                "缺口字段": display(lead.get("数据缺口")),
                "建议负责人": display(lead.get("负责人")),
                "状态": "待处理",
                "截止建议": (TODAY + timedelta(days=7)).strftime("%Y-%m-%d"),
                "关联商圈": display(lead.get("关联商圈")),
                "备注": "铺位线索不是正式推荐，现场核验后再进入审批。",
                "数据更新时间": TODAY_TEXT,
                "来源等级": display(lead.get("来源等级")),
                "数据质量评分": display(lead.get("数据质量评分")),
                "质量门槛说明": display(lead.get("质量门槛说明")),
                "任务状态分组": "本周重点",
                "关联铺位线索": display(lead.get("线索ID")),
            }
        )
        task_no += 1
    headers = [text(cell.value) for cell in ws[1]]
    rows = old_rows + generated
    rewrite_sheet(wb, "核验任务表", headers, rows)


def enhance_reports(wb, all_records, site_leads):
    if "报告输出表" not in wb.sheetnames:
        return
    headers, existing = sheet_records(wb["报告输出表"])
    rows = [row for _, row in existing if not text(row.get("报告ID")).startswith("REPORT-V05-")]
    fuzhou_leads = [row for row in site_leads if text(row.get("城市")) == "福州"][:12]
    fuzhou_areas = sorted([row for _, row in all_records.get("商圈库", []) if text(row.get("城市")) == "福州"], key=lambda row: -row_score(row))[:8]
    rows.extend(
        [
            {
                "报告ID": "REPORT-V05-WEEKLY-FUZHOU",
                "对象类型": "核验清单",
                "对象ID": "FUZHOU-WEEKLY",
                "标题": "福州本周重点核验清单",
                "结论": f"本周优先核验{len(fuzhou_areas)}个商圈和{len(fuzhou_leads)}条铺位线索，所有质量分低于70的对象只显示潜力推荐/待核验。",
                "地图摘要": "从福建省图进入福州，下钻鼓楼、台江、仓山、晋安等区县，重点查看商圈点位、半径圈和任务密度。",
                "评分摘要": "排序依据为城市/商圈评分、数据质量、竞品/租金/照片缺口和店型匹配。",
                "证据摘要": "现阶段公开资料和AI线索较多，平台截图、租金报价、实地照片仍是核心缺口。",
                "风险摘要": "不得把低质量线索作为正式推荐；需防止高分商圈因租金和竞品压力导致回本困难。",
                "下一步动作": "拓展经理按P1任务走访，选址分析师回写截图/照片/租金并重算评分。",
                "生成状态": "V0.5可预览",
                "数据更新时间": TODAY_TEXT,
            },
            {
                "报告ID": "REPORT-V05-MONTHLY-FUJIAN",
                "对象类型": "管理层摘要",
                "对象ID": "FUJIAN-MONTHLY",
                "标题": "福建选址月度管理层摘要",
                "结论": "福建优先保持福州样板深挖，厦门、泉州同步增强，其他城市保持骨架和数据缺口看板。",
                "地图摘要": "省级地图支持城市优先级、商圈数、竞品密度、数据质量和任务密度图层切换。",
                "评分摘要": "V0.5引入L0-L5来源等级和质量门槛，低于70分不允许直接推荐。",
                "证据摘要": "城市和区县基础盘逐步完善，商圈与铺位级仍需平台、租金和现场证据。",
                "风险摘要": "数据质量不足时，必须先安排核验任务，不能直接进入加盟/开店承诺。",
                "下一步动作": "每周更新核验中心，每月复盘评分权重和已开门店经营反馈。",
                "生成状态": "V0.5可预览",
                "数据更新时间": TODAY_TEXT,
            },
        ]
    )
    rewrite_sheet(wb, "报告输出表", headers, rows)


def write_source_level_sheet(wb):
    rows = [
        {
            "等级": code,
            "标签": label,
            "权重": weight,
            "典型来源": source,
            "使用说明": note,
            "是否可直接推荐": "是" if code in {"L4", "L5"} else ("需配合其他证据" if code == "L3" else "否"),
        }
        for code, label, weight, source, note in SOURCE_LEVELS
    ]
    rewrite_sheet(wb, "数据源等级表", ["等级", "标签", "权重", "典型来源", "使用说明", "是否可直接推荐"], rows)


def main():
    workbook = find_workbook()
    backup = workbook.with_name(f"{workbook.stem}.bak-v05-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook.suffix}")
    shutil.copy2(workbook, backup)
    wb = load_workbook(workbook)
    headers_by_sheet, all_records = upgrade_existing_sheets(wb)
    upgrade_candidates(wb, all_records)
    headers_by_sheet, all_records = refresh_records(wb)
    matrix_rows = build_field_matrix(all_records)
    site_leads = build_site_leads(all_records)
    city_kpi, district_kpi = build_kpis(all_records, matrix_rows)
    business_profiles = build_business_profiles(all_records, matrix_rows)
    street_environment = build_street_environment(all_records)

    write_source_level_sheet(wb)
    rewrite_sheet(
        wb,
        "字段核验矩阵",
        ["矩阵ID", "对象类型", "对象ID", "省份", "城市", "区县", "名称", "字段", "当前值", "是否缺失", "来源等级", "字段权重", "核验优先级", "核验动作", "数据更新时间"],
        matrix_rows,
    )
    rewrite_sheet(
        wb,
        "城市KPI表",
        ["省份", "城市", "推荐等级", "综合评分", "严格决策结论", "来源等级", "平均数据质量分", "区县数", "商圈数", "街道数", "POI数", "竞品数", "租金样本数", "核验任务数", "P1任务数", "推荐对象数", "潜力推荐数", "缺竞品", "缺租金", "缺照片", "缺平台数据", "数据缺口摘要", "下一步动作", "数据更新时间"],
        city_kpi,
    )
    rewrite_sheet(
        wb,
        "区县KPI表",
        ["省份", "城市", "区县", "推荐等级", "综合评分", "严格决策结论", "来源等级", "平均数据质量分", "商圈数", "街道数", "POI数", "竞品数", "租金样本数", "核验任务数", "P1任务数", "缺竞品", "缺租金", "缺照片", "缺平台数据", "数据缺口摘要", "下一步动作", "数据更新时间"],
        district_kpi,
    )
    rewrite_sheet(
        wb,
        "商圈画像表",
        ["画像ID", "省份", "城市", "区县", "街道", "商圈名称", "商圈类型", "推荐等级", "综合评分", "严格决策结论", "数据质量评分", "来源等级", "质量门槛说明", "1.5公里POI数", "3公里POI数", "3公里竞品数", "3公里租金样本数", "竞品核验数", "租金样本数", "照片证据数", "缺竞品", "缺租金", "缺照片", "缺平台数据", "主要依据", "主要风险", "下一步动作", "数据缺口", "数据更新时间"],
        business_profiles,
    )
    rewrite_sheet(
        wb,
        "街道环境表",
        ["环境ID", "省份", "城市", "区县", "街道/片区", "关联商圈", "推荐等级", "环境评分", "严格决策结论", "数据质量评分", "来源等级", "质量门槛说明", "人流/可见性", "餐饮集聚", "住宅", "办公", "停车", "竞品", "租金", "堂食判断", "外卖判断", "数据缺口", "下一步动作", "数据更新时间"],
        street_environment,
    )
    rewrite_sheet(
        wb,
        "铺位线索表",
        ["线索ID", "来源对象ID", "省份", "城市", "区县", "街道/片区", "关联商圈", "铺位线索名称", "线索类型", "店型模式", "经度", "纬度", "面积建议", "租金/招商状态", "楼层/门头", "推荐结论", "推荐等级", "综合评分", "数据质量评分", "来源等级", "质量门槛说明", "数据缺口", "核验任务", "负责人", "状态", "下一步动作", "图片要求", "数据更新时间"],
        site_leads,
    )

    headers_by_sheet, all_records = refresh_records(wb)
    enhance_tasks(wb, all_records, site_leads)
    headers_by_sheet, all_records = refresh_records(wb)
    enhance_reports(wb, all_records, site_leads)

    wb.save(workbook)
    print(f"V0.5 workbook upgraded: {workbook}")
    print(f"Backup created: {backup}")
    print(f"Field matrix rows: {len(matrix_rows)}")
    print(f"Site leads: {len(site_leads)}")


if __name__ == "__main__":
    main()
