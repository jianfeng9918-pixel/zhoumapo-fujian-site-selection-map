from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


REQUIRED_SHEETS = {"城市库", "商圈库", "街道库", "数据源登记表"}
TODAY = "2026-05-10"


def find_workbook() -> Path:
    candidates = []
    for path in Path("C:/CodexData").rglob("*.xlsx"):
        if ".bak-" in path.name:
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if REQUIRED_SHEETS.issubset(set(wb.sheetnames)):
                bonus = sum(1 for sheet in ["街道决策表", "门店分布表", "竞品品牌表", "需求开发库"] if sheet in wb.sheetnames)
                candidates.append((bonus, path.stat().st_mtime, path))
        except Exception:
            continue
    if candidates:
        return sorted(candidates, reverse=True)[0][2]
    raise FileNotFoundError("未找到周麻婆选址数据总表")


def text(value, default="待补"):
    if value is None:
        return default
    value = str(value).strip()
    return value if value else default


def to_float(value, default=0.0):
    value = text(value, "")
    keep = "".join(ch for ch in value if ch.isdigit() or ch == ".")
    try:
        return float(keep) if keep else default
    except Exception:
        return default


def records(wb, sheet):
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [text(cell.value, "") for cell in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {header: value for header, value in zip(headers, raw) if header}
        if any(text(value, "") for value in row.values()):
            rows.append(row)
    return rows


def pick(row, *keys, default="待补"):
    for key in keys:
        if key in row and text(row.get(key), ""):
            return text(row.get(key))
    return default


def grade_from_score(score):
    if score >= 85:
        return "A"
    if score >= 75:
        return "B"
    if score >= 65:
        return "C"
    if score > 0:
        return "D"
    return "待定"


def store_type(row):
    blob = " ".join(
        pick(row, key, default="")
        for key in ["店型匹配", "关联商圈", "街道/片区", "商圈类型", "住宅", "办公", "备注"]
    )
    types = []
    if any(token in blob for token in ["住宅", "社区", "底商", "小区"]):
        types.append("社区店")
    if any(token in blob for token in ["万达", "商场", "购物", "商业体", "广场", "Mall", "mall"]):
        types.append("购物中心店")
    if any(token in blob for token in ["高校", "大学", "学校", "办公", "CBD", "写字楼"]):
        types.append("高校/办公店")
    if any(token in blob for token in ["商圈", "街区", "步行街", "夜市", "文旅"]):
        types.append("商圈店")
    if not types:
        types.append("商圈店")
    return "、".join(dict.fromkeys(types))


def capacity(score, quality, own_count, competitor_count):
    if quality < 55:
        return "0-1个待核验"
    if score >= 86 and own_count == 0:
        return "2-3个优先深挖"
    if score >= 78:
        return "1-2个优先核验"
    if score >= 68:
        return "1个谨慎观察"
    return "暂不进入本周"


def decision(score, quality):
    if quality < 55:
        return "先补数据"
    if quality < 70:
        return "潜力推荐/待核验" if score >= 78 else "谨慎/待核验"
    if score >= 85:
        return "推荐"
    if score >= 75:
        return "谨慎推荐"
    return "谨慎"


def match_rows(rows, city, district="", keyword=""):
    result = []
    for row in rows:
        if text(row.get("城市"), "") != city:
            continue
        if district and text(row.get("区县"), "") not in {"", "待补", district}:
            continue
        haystack = " ".join(text(row.get(key), "") for key in ["名称", "关联商圈", "街道/片区", "门店名称", "地址"])
        if keyword and keyword not in haystack:
            continue
        result.append(row)
    return result


def build_street_decisions(wb):
    streets = records(wb, "街道库")
    env_rows = records(wb, "街道环境表")
    profiles = records(wb, "商圈画像表")
    competitors = records(wb, "竞品门店库")
    rents = records(wb, "租金样本库")
    tasks = records(wb, "核验任务表")
    env_by_name = {(pick(row, "城市", default=""), pick(row, "街道/片区", default="")): row for row in env_rows}
    profile_by_area = {(pick(row, "城市", default=""), pick(row, "商圈名称", default="")): row for row in profiles}
    rows = []
    for index, street in enumerate(streets, 1):
        city = pick(street, "城市")
        district = pick(street, "区县")
        street_name = pick(street, "街道/片区", "街道", "名称")
        area = pick(street, "关联商圈")
        env = env_by_name.get((city, street_name), {})
        profile = profile_by_area.get((city, area), {})
        score = to_float(pick(env, "环境评分", default=pick(street, "环境评分", "综合评分", default="0")), 0)
        if not score:
            score = to_float(pick(street, "综合评分", "环境评分", default="0"), 0)
        quality = to_float(pick(env, "数据质量评分", default=pick(street, "数据质量评分", default="0")), 0)
        comp = match_rows(competitors, city, district, area or street_name)
        rent = match_rows(rents, city, district, area or street_name)
        task_count = len(match_rows(tasks, city, district, area or street_name))
        own_count = 0
        competitor_count = len(comp)
        grade = pick(env, "推荐等级", default=pick(street, "最高推荐等级", "推荐等级", default=grade_from_score(score)))
        fit = store_type({**street, **env, **profile})
        row_decision = decision(score, quality)
        row_capacity = capacity(score, quality, own_count, competitor_count)
        reason = pick(
            env,
            "主要依据",
            default=pick(street, "主要依据", default=pick(profile, "主要依据", default="待补")),
        )
        risk = pick(
            env,
            "主要风险",
            default=pick(street, "主要风险", default=pick(profile, "主要风险", default="待补")),
        )
        gap = pick(
            env,
            "数据缺口",
            default=pick(street, "数据缺口", default=pick(profile, "数据缺口", default="待补")),
        )
        next_action = pick(
            env,
            "下一步动作",
            default="用美团选址工具核验竞品评分/销量，用招商或实地走访核验租金、门头、动线和合同条件",
        )
        rows.append(
            {
                "决策ID": f"STDEC-V07-{index:04d}",
                "省份": pick(street, "省份", default="福建"),
                "城市": city,
                "区县": district,
                "街道/片区": street_name,
                "关联商圈": area,
                "经度": pick(street, "经度", default=pick(env, "经度", default="待补")),
                "纬度": pick(street, "纬度", default=pick(env, "纬度", default="待补")),
                "推荐等级": grade,
                "街道评分": round(score, 1) if score else "待补",
                "数据质量评分": round(quality, 1) if quality else "待补",
                "严格决策结论": row_decision,
                "适合店型": fit,
                "可开店容量": row_capacity,
                "常住人口/覆盖人口": pick(street, "常住人口", "覆盖人口", default="待补"),
                "住宅成熟度": pick(env, "住宅", default=pick(street, "住宅", default="待补")),
                "学校/家庭客群": pick(street, "学校/家庭客群", default="待补：需高德/美团/实地核验学校与家庭客群"),
                "办公/商业配套": pick(env, "办公", default=pick(street, "办公", default="待补")),
                "餐饮业态": pick(env, "餐饮集聚", default=pick(street, "餐饮集聚", default="待补：需补30-70元中式/川菜/家常菜餐饮密度")),
                "租金压力": pick(env, "租金", default=pick(profile, "租金压力", default="待补")),
                "竞品压力": pick(env, "竞品", default=f"{competitor_count}条竞品线索，待平台核验"),
                "周麻婆现有门店数": own_count,
                "友商/竞品门店数": competitor_count,
                "租金样本数": len(rent),
                "核验任务数": task_count,
                "主要依据": reason,
                "主要风险": risk,
                "数据缺口": gap,
                "下一步核验动作": next_action,
                "美团核验重点": "评分、评论数、人均、月销量/热度、榜单、附近流水指数、同价位餐饮表现",
                "实地核验重点": "门头可见度、停车、动线、台阶/转角、周边住宅入住率、餐饮业态、租金合同条件",
                "来源等级": pick(env, "来源等级", default=pick(street, "来源等级", default="L1")),
                "数据更新时间": TODAY,
            }
        )
    rows.sort(key=lambda r: (r["城市"] != "福州", -to_float(r["街道评分"]), r["城市"], r["区县"], r["街道/片区"]))
    return rows


def store_distribution_template():
    return [
        {
            "门店ID": "STORE-TEMPLATE-001",
            "品牌": "周麻婆",
            "门店名称": "待导入门店名称",
            "省份": "福建",
            "城市": "待补",
            "区县": "待补",
            "街道/片区": "待补",
            "地址": "待导入完整地址",
            "经度": "待补",
            "纬度": "待补",
            "营业状态": "待导入",
            "店型": "待补",
            "来源": "用户提供门店地址表",
            "来源等级": "L3",
            "备注": "导入周麻婆140家门店地址后，用于现有门店图层和空白机会判断。",
            "数据更新时间": TODAY,
        }
    ]


def competitor_brands():
    brands = [
        ("COMP-BRAND-001", "小叫天", "福建重点友商", "中式快餐/家常菜", "30-70元", "福州高集中，需要重点看空白商圈和贴身竞争"),
        ("COMP-BRAND-002", "最得意", "福建重点友商", "川湘/家常菜", "30-70元", "用于福州、泉州、厦门竞品密度判断"),
        ("COMP-BRAND-003", "四方桌", "福建重点友商", "家常菜/聚餐", "40-80元", "同事提到老板关注，可纳入品牌对标"),
        ("COMP-BRAND-004", "大丰收", "扩展竞品", "家常菜/中餐", "40-90元", "部分门店做家常菜，可作为同价位业态观察"),
        ("COMP-BRAND-005", "美果", "待确认竞品", "中餐/区域品牌", "待补", "先收集，后续按价格带和客群判断是否纳入核心竞品"),
        ("COMP-BRAND-006", "姑奶奶", "跨区竞品", "中式家常菜/区域连锁", "30-70元", "宁德、闽南、广东等区域曾与周麻婆竞争，可纳入参考"),
    ]
    return [
        {
            "品牌ID": brand_id,
            "品牌名称": name,
            "竞品类型": comp_type,
            "餐饮业态": category,
            "价格带": price,
            "重点说明": note,
            "是否核心竞品": "是" if index <= 3 else "待确认",
            "门店数据状态": "待导入/待平台核验",
            "来源": "2026-05-10同事讨论纪要",
            "来源等级": "L2",
            "核验动作": "补品牌门店名、地址、评分、人均、销量/热度、来源截图",
            "数据更新时间": TODAY,
        }
        for index, (brand_id, name, comp_type, category, price, note) in enumerate(brands, 1)
    ]


def requirement_backlog():
    data = [
        ("REQ-V07-001", "页面主线", "首页聚焦城市地图+街道决策榜，POI和报告放高级入口", "P1", "本次实现"),
        ("REQ-V07-002", "地图交互", "保留两级地图，上方福建，下方城市/区县/街道，不做复杂三级跳转", "P1", "本次实现"),
        ("REQ-V07-003", "返回体验", "点击城市后保留面包屑和原地图，减少来回跳转", "P1", "本次实现"),
        ("REQ-V07-004", "图层", "支持周麻婆门店、竞品门店、推荐商圈、推荐街道、空白机会、数据质量开关", "P1", "本次实现框架"),
        ("REQ-V07-005", "门店数据", "导入周麻婆100+门店地址，用于空白机会和已有门店密度", "P1", "待用户提供地址表"),
        ("REQ-V07-006", "竞品数据", "收集小叫天、最得意、四方桌、大丰收、美果、姑奶奶门店名/评分/地址/销量", "P1", "待平台截图或人工导出"),
        ("REQ-V07-007", "街道指标", "增加常住人口、平均房价、入住率、学校、住宅、办公、餐饮业态、租金", "P1", "本次建字段"),
        ("REQ-V07-008", "餐饮业态", "不只看竞品，要看30-70元中式/川菜/家常菜餐饮密度", "P1", "本次建字段"),
        ("REQ-V07-009", "美团工具", "本系统做前置筛选，美团选址工具做最终点位核验", "P1", "定位确认"),
        ("REQ-V07-010", "导出表格", "底部数据可导出为表格，便于同事继续加工", "P2", "后续实现"),
        ("REQ-V07-011", "使用对象", "内部给老板/拓展/选址看，不做加盟商外宣版", "P1", "定位确认"),
        ("REQ-V07-012", "高级模块", "POI、报告、字段矩阵暂不删，收进高级数据", "P1", "本次实现"),
    ]
    return [
        {
            "需求ID": req_id,
            "需求类别": category,
            "需求内容": content,
            "优先级": priority,
            "处理状态": status,
            "来源": "周麻婆选址决策平台规划.txt",
            "暂缓原因": "待数据源或后续版本" if "待" in status else "不暂缓",
            "下一步动作": "按优先级进入V0.7/V0.8迭代",
            "数据更新时间": TODAY,
        }
        for req_id, category, content, priority, status in data
    ]


def rewrite_sheet(wb, title, headers, rows):
    if title in wb.sheetnames:
        idx = wb.sheetnames.index(title)
        del wb[title]
        ws = wb.create_sheet(title, idx)
    else:
        ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "待补") for header in headers])
    style_sheet(ws)


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="EAF3EF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="173328")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        header = text(ws.cell(1, col[0].column).value, "")
        width = 14
        if header in {"主要依据", "主要风险", "数据缺口", "下一步核验动作", "美团核验重点", "实地核验重点", "重点说明", "需求内容"}:
            width = 36
        elif len(header) >= 6:
            width = 18
        ws.column_dimensions[col[0].column_letter].width = width
        for cell in col:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    workbook = find_workbook()
    backup = workbook.with_name(f"{workbook.stem}.bak-v07-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook.suffix}")
    shutil.copy2(workbook, backup)
    wb = load_workbook(workbook)

    street_rows = build_street_decisions(wb)
    street_headers = [
        "决策ID",
        "省份",
        "城市",
        "区县",
        "街道/片区",
        "关联商圈",
        "经度",
        "纬度",
        "推荐等级",
        "街道评分",
        "数据质量评分",
        "严格决策结论",
        "适合店型",
        "可开店容量",
        "常住人口/覆盖人口",
        "住宅成熟度",
        "学校/家庭客群",
        "办公/商业配套",
        "餐饮业态",
        "租金压力",
        "竞品压力",
        "周麻婆现有门店数",
        "友商/竞品门店数",
        "租金样本数",
        "核验任务数",
        "主要依据",
        "主要风险",
        "数据缺口",
        "下一步核验动作",
        "美团核验重点",
        "实地核验重点",
        "来源等级",
        "数据更新时间",
    ]
    rewrite_sheet(wb, "街道决策表", street_headers, street_rows)

    store_headers = ["门店ID", "品牌", "门店名称", "省份", "城市", "区县", "街道/片区", "地址", "经度", "纬度", "营业状态", "店型", "来源", "来源等级", "备注", "数据更新时间"]
    rewrite_sheet(wb, "门店分布表", store_headers, store_distribution_template())

    brand_headers = ["品牌ID", "品牌名称", "竞品类型", "餐饮业态", "价格带", "重点说明", "是否核心竞品", "门店数据状态", "来源", "来源等级", "核验动作", "数据更新时间"]
    rewrite_sheet(wb, "竞品品牌表", brand_headers, competitor_brands())

    req_headers = ["需求ID", "需求类别", "需求内容", "优先级", "处理状态", "来源", "暂缓原因", "下一步动作", "数据更新时间"]
    rewrite_sheet(wb, "需求开发库", req_headers, requirement_backlog())

    wb.save(workbook)
    counts = Counter(row["城市"] for row in street_rows)
    print(f"V0.7 workbook upgraded: {workbook}")
    print(f"Backup created: {backup}")
    print(f"Street decisions: {len(street_rows)}")
    print(dict(counts.most_common(6)))


if __name__ == "__main__":
    main()
