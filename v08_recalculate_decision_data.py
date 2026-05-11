from __future__ import annotations

import math
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "01-选址数据总表" / "周麻婆选址数据总表.xlsx"
TODAY = datetime.now().strftime("%Y-%m-%d")

FUJIAN = "福建"
FUJIAN_CITIES = ["福州", "厦门", "泉州", "漳州", "莆田", "三明", "南平", "龙岩", "宁德"]
KEY_CITIES = {"福州", "厦门", "泉州"}
CORE_COMPETITORS = {"小叫天", "醉得意", "四方桌", "大丰收", "姑奶奶"}
MISSING_VALUES = {"", "待补", "-", "None", "无", "待识别", "待导入", "待算"}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def is_missing(value) -> bool:
    text = clean(value)
    return text in MISSING_VALUES or text.startswith("待补")


def to_float(value, default: float | None = None):
    text = clean(value)
    if not text:
        return default
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default
    try:
        return float(match.group())
    except ValueError:
        return default


def to_int(value, default=0) -> int:
    parsed = to_float(value, None)
    return int(parsed) if parsed is not None else default


def normalize(value) -> str:
    text = clean(value)
    text = re.sub(r"[（）()·\-\s/、,，]+", "", text)
    for token in ["街道", "片区", "商圈", "周边", "待核验", "门店", "店", "广场", "中心"]:
        text = text.replace(token, "")
    return text


def has_geo(row) -> bool:
    return to_float(row.get("经度"), None) is not None and to_float(row.get("纬度"), None) is not None


def haversine_km(a_lng, a_lat, b_lng, b_lat) -> float | None:
    vals = [to_float(v, None) for v in [a_lng, a_lat, b_lng, b_lat]]
    if any(v is None for v in vals):
        return None
    lng1, lat1, lng2, lat2 = [math.radians(v) for v in vals]
    dlng = lng2 - lng1
    dlat = lat2 - lat1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
    return 6371 * 2 * math.asin(math.sqrt(h))


def workbook_path() -> Path:
    if WORKBOOK.exists():
        return WORKBOOK
    candidates = [path for path in ROOT.rglob("*.xlsx") if ".bak-" not in path.name and path.name != "brand_competitors.xlsx"]
    if not candidates:
        raise FileNotFoundError("未找到周麻婆选址数据总表")
    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def headers(ws) -> list[str]:
    return [clean(cell.value) for cell in ws[1]]


def ensure_headers(ws, names: list[str]):
    current = headers(ws)
    for name in names:
        if name not in current:
            ws.cell(1, len(current) + 1, name)
            current.append(name)


def sheet_records(ws) -> list[dict[str, str]]:
    h = headers(ws)
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row = {header: clean(ws.cell(row_num, col_num).value) for col_num, header in enumerate(h, start=1)}
        if any(clean(value) for value in row.values()):
            row["_row"] = row_num
            rows.append(row)
    return rows


def set_cell(ws, row_num: int, header: str, value):
    h = headers(ws)
    if header not in h:
        ws.cell(1, len(h) + 1, header)
        h.append(header)
    ws.cell(row_num, h.index(header) + 1, value)


def replace_rows(ws, rows: list[dict[str, str]]):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    h = headers(ws)
    for row in rows:
        ws.append([row.get(header, "") for header in h])


def same_place(row, decision, max_km=1.35) -> bool:
    if clean(row.get("城市")) != clean(decision.get("城市")):
        return False
    row_district = clean(row.get("区县"))
    decision_district = clean(decision.get("区县"))
    if row_district and decision_district and row_district != decision_district:
        return False
    row_names = [normalize(row.get("街道/片区")), normalize(row.get("关联商圈")), normalize(row.get("门店名称")), normalize(row.get("名称"))]
    decision_names = [normalize(decision.get("街道/片区")), normalize(decision.get("关联商圈"))]
    for a in row_names:
        for b in decision_names:
            if a and b and (a == b or a in b or b in a):
                return True
    distance = haversine_km(row.get("经度"), row.get("纬度"), decision.get("经度"), decision.get("纬度"))
    return distance is not None and distance <= max_km


def same_area(row, profile) -> bool:
    if clean(row.get("城市")) != clean(profile.get("城市")):
        return False
    if clean(row.get("区县")) and clean(profile.get("区县")) and clean(row.get("区县")) != clean(profile.get("区县")):
        return False
    row_names = [normalize(row.get("街道/片区")), normalize(row.get("关联商圈")), normalize(row.get("门店名称")), normalize(row.get("名称"))]
    profile_names = [normalize(profile.get("商圈名称")), normalize(profile.get("街道"))]
    return any(a and b and (a == b or a in b or b in a) for a in row_names for b in profile_names)


def city_filter(rows):
    return [row for row in rows if clean(row.get("省份")) == FUJIAN or clean(row.get("城市")) in FUJIAN_CITIES]


def fujian_rows(rows):
    return [row for row in rows if clean(row.get("省份")) == FUJIAN]


def source_level_score(level: str) -> int:
    return {"L5": 20, "L4": 18, "L3": 15, "L2": 10, "L1": 6, "L0": 0}.get(clean(level), 0)


def calc_street_quality(row, own_count, comp_count, core_comp_count, rent_count, poi_count, gaps) -> float:
    fields = [
        row.get("经度"),
        row.get("纬度"),
        row.get("常住人口/覆盖人口"),
        row.get("住宅成熟度"),
        row.get("学校/家庭客群"),
        row.get("办公/商业配套"),
        row.get("餐饮业态"),
        row.get("租金压力"),
        row.get("竞品压力"),
        own_count,
        comp_count,
        rent_count,
        poi_count,
    ]
    filled = sum(1 for value in fields if not is_missing(value))
    coverage = filled / len(fields)
    platform_bonus = min(core_comp_count, 5) * 1.6 + min(rent_count, 3) * 2 + min(poi_count, 5)
    quality = 42 + coverage * 36 + source_level_score(row.get("来源等级")) + platform_bonus
    if gaps:
        quality -= min(12, len(gaps) * 2)
    return round(max(0, min(95, quality)), 1)


def tags_for(row, own_count, core_comp_count, rent_count, poi_count) -> list[str]:
    score = to_float(row.get("街道评分"), 0) or 0
    tags = []
    if own_count == 0 and score >= 82:
        tags.append("空白机会")
    if own_count == 0 and core_comp_count >= 3:
        tags.append("竞争验证强")
    if own_count > 0:
        tags.append("本店已覆盖")
    if own_count >= 2:
        tags.append("门店过密/谨慎")
    if rent_count == 0:
        tags.append("需租金核验")
    if poi_count < 3:
        tags.append("需POI核验")
    return tags or ["常规跟踪"]


def capacity_for(score, own_count, core_comp_count):
    if own_count >= 2:
        return "暂缓新增/防内耗"
    if own_count == 1 and score >= 86 and core_comp_count >= 4:
        return "可评估1家补位"
    if own_count == 0 and score >= 86:
        return "可优先评估1-2家"
    if own_count == 0 and score >= 76:
        return "可评估1家"
    return "暂不优先"


def strict_decision(score, quality, own_count, tags):
    if own_count >= 2:
        return "谨慎/防内耗"
    if quality < 70:
        return "潜力推荐/待核验" if score >= 82 else "先补数据"
    if "空白机会" in tags or "竞争验证强" in tags:
        return "潜力推荐/待核验"
    if score >= 76:
        return "谨慎/待核验"
    return "先补数据"


def update_street_decisions(wb):
    street_ws = wb["街道决策表"]
    ensure_headers(
        street_ws,
        ["机会标签", "空白机会类型", "五大竞品门店数", "同区县周麻婆门店数", "同区县竞品门店数", "V0.8本周核验建议"],
    )
    streets = fujian_rows(sheet_records(street_ws))
    own_rows = fujian_rows(sheet_records(wb["门店分布表"]))
    comp_rows = fujian_rows(sheet_records(wb["竞品门店库"]))
    rent_rows = fujian_rows(sheet_records(wb["租金样本库"]))
    poi_rows = fujian_rows(sheet_records(wb["POI库"]))

    for row in streets:
        own_matches = [item for item in own_rows if same_place(item, row)]
        comp_matches = [item for item in comp_rows if same_place(item, row)]
        core_matches = [item for item in comp_matches if clean(item.get("竞品品牌")) in CORE_COMPETITORS]
        rent_matches = [item for item in rent_rows if same_place(item, row, max_km=1.5)]
        poi_matches = [item for item in poi_rows if same_place(item, row, max_km=1.5)]
        district_own = [item for item in own_rows if clean(item.get("城市")) == row.get("城市") and clean(item.get("区县")) == row.get("区县")]
        district_comp = [item for item in comp_rows if clean(item.get("城市")) == row.get("城市") and clean(item.get("区县")) == row.get("区县")]

        own_count = len(own_matches)
        comp_count = len(comp_matches)
        core_count = len(core_matches)
        rent_count = len(rent_matches)
        poi_count = len(poi_matches)
        score = to_float(row.get("街道评分"), 0) or 0
        tags = tags_for(row, own_count, core_count, rent_count, poi_count)
        gaps = []
        if rent_count == 0:
            gaps.append("缺租金/招商报价")
        if poi_count < 3:
            gaps.append("缺POI客群密度")
        if core_count > 0:
            gaps.append("需五大竞品截图留档")
        gaps.append("缺实地照片/门头动线")
        if not has_geo(row):
            gaps.append("缺地理编码")
        quality = calc_street_quality(row, own_count, comp_count, core_count, rent_count, poi_count, gaps)
        capacity = capacity_for(score, own_count, core_count)
        decision = strict_decision(score, quality, own_count, tags)
        opportunity_type = "高潜空白区" if "空白机会" in tags else ("竞争验证区" if "竞争验证强" in tags else ("已覆盖维护区" if own_count else "常规观察区"))
        task_count = 1 if any(tag in tags for tag in ["空白机会", "竞争验证强", "需租金核验"]) else 0

        basis = clean(row.get("主要依据"))
        basis_add = f"V0.8重算：本街道/片区匹配周麻婆{own_count}家、五大竞品{core_count}家、全部竞品{comp_count}家、租金样本{rent_count}条、POI线索{poi_count}条"
        risk = clean(row.get("主要风险"))
        risk_add = f"同区县周麻婆{len(district_own)}家、竞品{len(district_comp)}家；街道归属仍需用地图/实地抽样核验"
        next_action = "本周优先核验：美团/点评截图、竞品营业状态、招商租金、门头动线、停车和周边100米环境"

        updates = {
            "周麻婆现有门店数": own_count,
            "友商/竞品门店数": comp_count,
            "五大竞品门店数": core_count,
            "同区县周麻婆门店数": len(district_own),
            "同区县竞品门店数": len(district_comp),
            "租金样本数": rent_count,
            "核验任务数": task_count,
            "机会标签": "、".join(tags),
            "空白机会类型": opportunity_type,
            "可开店容量": capacity,
            "数据质量评分": quality,
            "严格决策结论": decision,
            "竞品压力": "高" if core_count >= 6 else ("中高" if core_count >= 3 else ("中" if comp_count else "低/待核验")),
            "数据缺口": "；".join(gaps) if gaps else "核心字段已覆盖，进入抽样复核",
            "主要依据": basis_add if basis_add in basis else (f"{basis}；{basis_add}" if basis else basis_add),
            "主要风险": risk_add if risk_add in risk else (f"{risk}；{risk_add}" if risk else risk_add),
            "下一步核验动作": next_action,
            "美团核验重点": "五大竞品评分、评论数、人均、销量/热度、榜单、营业状态截图",
            "实地核验重点": "门头可见度、动线、停车、外摆、竞品排队、客群时段、租金合同",
            "来源等级": "L3" if core_count or own_count else clean(row.get("来源等级")) or "L2",
            "V0.8本周核验建议": next_action if row.get("城市") in KEY_CITIES and task_count else "纳入月度补数",
            "数据更新时间": TODAY,
        }
        for key, value in updates.items():
            set_cell(street_ws, row["_row"], key, value)


def rebuild_field_matrix(wb):
    matrix_ws = wb["字段核验矩阵"]
    core_specs = [
        ("城市", wb["城市库"], ["推荐等级", "综合评分", "经度", "纬度", "常住人口", "GDP/社零/第三产业", "来源等级", "数据更新时间"], "记录ID", "名称"),
        ("商圈", wb["商圈库"], ["名称", "城市", "区县", "经度", "纬度", "推荐等级", "综合评分", "商圈类型", "主要依据", "主要风险", "数据缺口"], "记录ID", "名称"),
        ("街道决策", wb["街道决策表"], ["经度", "纬度", "推荐等级", "街道评分", "数据质量评分", "适合店型", "可开店容量", "机会标签", "周麻婆现有门店数", "五大竞品门店数", "友商/竞品门店数", "租金样本数", "数据缺口", "下一步核验动作"], "决策ID", "街道/片区"),
        ("周麻婆门店", wb["门店分布表"], ["门店名称", "城市", "区县", "街道/片区", "地址", "经度", "纬度", "营业状态", "店型"], "门店ID", "门店名称"),
        ("竞品门店", wb["竞品门店库"], ["门店名称", "竞品品牌", "城市", "区县", "街道/片区", "经度", "纬度", "评分", "评论数", "人均", "月销量/热度", "外卖表现", "堂食表现", "竞争压力", "验证状态"], "记录ID", "门店名称"),
        ("租金样本", wb["租金样本库"], ["城市", "区县", "街道/片区", "租金区间", "面积建议", "转让费", "物业费", "合同条件", "回本压力", "经度", "纬度"], "记录ID", "关联商圈"),
        ("POI", wb["POI库"], ["城市", "区县", "街道/片区", "名称", "POI类型", "经度", "纬度", "选址意义", "数据置信度"], "记录ID", "名称"),
        ("图片素材", wb["图片素材库"], ["城市", "区县", "关联对象", "图片类型", "图片路径/链接", "拍摄日期", "来源等级"], "记录ID", "关联对象"),
    ]
    rows = []
    index = 1
    critical = {"经度", "纬度", "评分", "评论数", "人均", "租金区间", "合同条件", "机会标签", "周麻婆现有门店数", "五大竞品门店数", "友商/竞品门店数"}
    for object_type, ws, fields, id_field, name_field in core_specs:
        for row in city_filter(sheet_records(ws)):
            if object_type in {"街道决策", "商圈", "城市"} and row.get("省份") != FUJIAN:
                continue
            for field in fields:
                value = row.get(field, "")
                missing = is_missing(value)
                priority = "P1" if missing and field in critical else ("P2" if missing else "普通")
                action = "补地理编码/经纬度" if field in {"经度", "纬度"} and missing else ("补字段并留来源截图/文件" if missing else "已覆盖，定期抽样复核")
                rows.append(
                    {
                        "矩阵ID": f"MATRIX-V08-{index:05d}",
                        "对象类型": object_type,
                        "对象ID": row.get(id_field) or row.get("记录ID") or row.get("门店ID") or f"{object_type}-{index}",
                        "省份": row.get("省份") or FUJIAN,
                        "城市": row.get("城市") or "待补",
                        "区县": row.get("区县") or "待补",
                        "名称": row.get(name_field) or row.get("名称") or row.get("门店名称") or row.get("街道/片区") or "待补",
                        "字段": field,
                        "当前值": value or "待补",
                        "是否缺失": "是" if missing else "否",
                        "来源等级": row.get("来源等级") or "L0",
                        "字段权重": "关键" if field in critical else "普通",
                        "核验优先级": priority,
                        "核验动作": action,
                        "数据更新时间": TODAY,
                    }
                )
                index += 1
    replace_rows(matrix_ws, rows)
    return rows


def update_kpis(wb):
    city_ws = wb["城市KPI表"]
    district_ws = wb["区县KPI表"]
    ensure_headers(city_ws, ["周麻婆门店数", "五大竞品门店数", "空白街道数", "高潜空白街道数"])
    ensure_headers(district_ws, ["周麻婆门店数", "五大竞品门店数", "空白街道数", "高潜空白街道数"])
    streets = fujian_rows(sheet_records(wb["街道决策表"]))
    own_rows = fujian_rows(sheet_records(wb["门店分布表"]))
    comp_rows = fujian_rows(sheet_records(wb["竞品门店库"]))
    areas = fujian_rows(sheet_records(wb["商圈库"]))
    poi_rows = fujian_rows(sheet_records(wb["POI库"]))
    rent_rows = fujian_rows(sheet_records(wb["租金样本库"]))
    tasks = fujian_rows(sheet_records(wb["核验任务表"]))
    candidates = fujian_rows(sheet_records(wb["候选清单表"]))

    def row_set(ws, row_num, values):
        for key, value in values.items():
            set_cell(ws, row_num, key, value)

    for row in sheet_records(city_ws):
        if clean(row.get("省份")) != FUJIAN:
            continue
        city = row.get("城市")
        city_streets = [item for item in streets if item.get("城市") == city]
        city_comp = [item for item in comp_rows if item.get("城市") == city]
        city_core = [item for item in city_comp if item.get("竞品品牌") in CORE_COMPETITORS]
        city_own = [item for item in own_rows if item.get("城市") == city]
        city_tasks = [item for item in tasks if item.get("城市") == city and item.get("状态") != "已完成"]
        qualities = [to_float(item.get("数据质量评分"), None) for item in city_streets]
        qualities = [value for value in qualities if value is not None]
        blank_streets = [item for item in city_streets if to_int(item.get("周麻婆现有门店数")) == 0]
        high_blank = [item for item in blank_streets if "空白机会" in clean(item.get("机会标签")) or "竞争验证强" in clean(item.get("机会标签"))]
        missing_rent = sum(1 for item in city_streets if to_int(item.get("租金样本数")) == 0)
        missing_comp = sum(1 for item in city_streets if to_int(item.get("五大竞品门店数")) == 0)
        row_set(
            city_ws,
            row["_row"],
            {
                "平均数据质量分": round(mean(qualities), 1) if qualities else 0,
                "商圈数": sum(1 for item in areas if item.get("城市") == city),
                "街道数": len(city_streets),
                "POI数": sum(1 for item in poi_rows if item.get("城市") == city),
                "竞品数": len(city_comp),
                "租金样本数": sum(1 for item in rent_rows if item.get("城市") == city),
                "核验任务数": len(city_tasks),
                "P1任务数": sum(1 for item in city_tasks if item.get("优先级") == "P1"),
                "推荐对象数": sum(1 for item in city_streets if item.get("严格决策结论") == "推荐"),
                "潜力推荐数": sum(1 for item in city_streets if "潜力推荐" in clean(item.get("严格决策结论"))),
                "缺竞品": missing_comp,
                "缺租金": missing_rent,
                "缺照片": len(city_streets),
                "缺平台数据": sum(1 for item in city_streets if to_int(item.get("五大竞品门店数")) == 0),
                "周麻婆门店数": len(city_own),
                "五大竞品门店数": len(city_core),
                "空白街道数": len(blank_streets),
                "高潜空白街道数": len(high_blank),
                "数据缺口摘要": f"缺租金街道{missing_rent}个；缺实地照片{len(city_streets)}个；五大竞品待补街道{missing_comp}个",
                "下一步动作": "优先核验高潜空白街道：租金、竞品截图、门头动线、停车和现场客流",
                "数据更新时间": TODAY,
            },
        )

    for row in sheet_records(district_ws):
        if clean(row.get("省份")) != FUJIAN:
            continue
        city = row.get("城市")
        district = row.get("区县")
        dist_streets = [item for item in streets if item.get("城市") == city and item.get("区县") == district]
        dist_comp = [item for item in comp_rows if item.get("城市") == city and item.get("区县") == district]
        dist_core = [item for item in dist_comp if item.get("竞品品牌") in CORE_COMPETITORS]
        dist_own = [item for item in own_rows if item.get("城市") == city and item.get("区县") == district]
        dist_tasks = [item for item in tasks if item.get("城市") == city and item.get("区县") == district and item.get("状态") != "已完成"]
        qualities = [to_float(item.get("数据质量评分"), None) for item in dist_streets]
        qualities = [value for value in qualities if value is not None]
        blank_streets = [item for item in dist_streets if to_int(item.get("周麻婆现有门店数")) == 0]
        high_blank = [item for item in blank_streets if "空白机会" in clean(item.get("机会标签")) or "竞争验证强" in clean(item.get("机会标签"))]
        row_set(
            district_ws,
            row["_row"],
            {
                "平均数据质量分": round(mean(qualities), 1) if qualities else row.get("平均数据质量分"),
                "商圈数": sum(1 for item in areas if item.get("城市") == city and item.get("区县") == district),
                "街道数": len(dist_streets),
                "POI数": sum(1 for item in poi_rows if item.get("城市") == city and item.get("区县") == district),
                "竞品数": len(dist_comp),
                "租金样本数": sum(1 for item in rent_rows if item.get("城市") == city and item.get("区县") == district),
                "核验任务数": len(dist_tasks),
                "P1任务数": sum(1 for item in dist_tasks if item.get("优先级") == "P1"),
                "缺竞品": sum(1 for item in dist_streets if to_int(item.get("五大竞品门店数")) == 0),
                "缺租金": sum(1 for item in dist_streets if to_int(item.get("租金样本数")) == 0),
                "缺照片": len(dist_streets),
                "缺平台数据": sum(1 for item in dist_streets if to_int(item.get("五大竞品门店数")) == 0),
                "周麻婆门店数": len(dist_own),
                "五大竞品门店数": len(dist_core),
                "空白街道数": len(blank_streets),
                "高潜空白街道数": len(high_blank),
                "数据缺口摘要": f"空白街道{len(blank_streets)}个；高潜空白{len(high_blank)}个；缺租金{sum(1 for item in dist_streets if to_int(item.get('租金样本数')) == 0)}个",
                "下一步动作": "按街道榜优先实地核验高潜空白区和竞争验证区",
                "数据更新时间": TODAY,
            },
        )


def update_profiles(wb):
    profile_ws = wb["商圈画像表"]
    profiles = fujian_rows(sheet_records(profile_ws))
    comp_rows = fujian_rows(sheet_records(wb["竞品门店库"]))
    rent_rows = fujian_rows(sheet_records(wb["租金样本库"]))
    poi_rows = fujian_rows(sheet_records(wb["POI库"]))
    for row in profiles:
        comp = [item for item in comp_rows if same_area(item, row)]
        core = [item for item in comp if item.get("竞品品牌") in CORE_COMPETITORS]
        rent = [item for item in rent_rows if same_area(item, row)]
        poi = [item for item in poi_rows if same_area(item, row)]
        gaps = []
        if not core:
            gaps.append("缺五大竞品截图")
        if not rent:
            gaps.append("缺租金/招商报价")
        gaps.append("缺实地照片")
        if len(poi) < 3:
            gaps.append("缺POI客群")
        quality = min(92, 45 + min(len(core), 6) * 3 + min(len(rent), 3) * 5 + min(len(poi), 5) * 2)
        for key, value in {
            "3公里竞品数": len(comp),
            "竞品核验数": len(core),
            "租金样本数": len(rent),
            "3公里租金样本数": len(rent),
            "1.5公里POI数": len(poi),
            "3公里POI数": len(poi),
            "缺竞品": 0 if core else 1,
            "缺租金": 0 if rent else 1,
            "缺照片": 1,
            "缺平台数据": 0 if core else 1,
            "数据质量评分": quality,
            "主要依据": f"V0.8重算：匹配五大竞品{len(core)}家、全部竞品{len(comp)}家、租金样本{len(rent)}条、POI线索{len(poi)}条",
            "主要风险": "商圈边界仍为文本匹配，需地图/实地复核",
            "下一步动作": "补商圈入口、主通道、竞品截图、招商租金和周边客群照片",
            "数据缺口": "；".join(gaps),
            "数据更新时间": TODAY,
        }.items():
            set_cell(profile_ws, row["_row"], key, value)


def rebuild_v08_tasks(wb):
    task_ws = wb["核验任务表"]
    id_col = headers(task_ws).index("任务ID") + 1
    for row_num in range(task_ws.max_row, 1, -1):
        if clean(task_ws.cell(row_num, id_col).value).startswith("TASK-V08-"):
            task_ws.delete_rows(row_num, 1)

    streets = fujian_rows(sheet_records(wb["街道决策表"]))
    selected = [
        row
        for row in streets
        if row.get("城市") in KEY_CITIES
        and any(tag in clean(row.get("机会标签")) for tag in ["空白机会", "竞争验证强", "需租金核验"])
        and "门店过密/谨慎" not in clean(row.get("机会标签"))
        and not ("本店已覆盖" in clean(row.get("机会标签")) and "空白机会" not in clean(row.get("机会标签")) and "竞争验证强" not in clean(row.get("机会标签")))
    ]
    selected = sorted(selected, key=lambda row: (row.get("城市") != "福州", -to_float(row.get("街道评分"), 0), -to_int(row.get("五大竞品门店数"))))[:40]
    h = headers(task_ws)
    for index, row in enumerate(selected, start=1):
        priority = "P1" if row.get("城市") == "福州" and to_float(row.get("街道评分"), 0) >= 82 else "P2"
        missing = clean(row.get("数据缺口"))
        task = {
            "任务ID": f"TASK-V08-STREET-{index:03d}",
            "对象类型": "街道决策",
            "对象ID": row.get("决策ID"),
            "省份": FUJIAN,
            "城市": row.get("城市"),
            "区县": row.get("区县"),
            "任务类型": "街道机会核验",
            "优先级": priority,
            "核验内容": "核验五大竞品截图、租金/招商报价、门头动线、停车、周边100米环境和实地照片",
            "缺口字段": missing,
            "建议负责人": "拓展/选址/加盟支持",
            "状态": "待处理",
            "截止建议": "本周" if priority == "P1" else "两周内",
            "关联商圈": row.get("关联商圈"),
            "备注": f"{row.get('机会标签')}；{row.get('可开店容量')}",
            "数据更新时间": TODAY,
            "来源等级": row.get("来源等级") or "L3",
            "数据质量评分": row.get("数据质量评分"),
            "质量门槛说明": "L3平台数据可用于前置筛选，不作为签约依据",
            "任务状态分组": "本周重点" if priority == "P1" else "近期补数",
            "关联铺位线索": "待补",
        }
        task_ws.append([task.get(header, "") for header in h])


def update_sources_and_log(wb, matrix_rows_count):
    source_ws = wb["数据源登记表"]
    log_ws = wb["更新日志"]
    def upsert(ws, id_header, values):
        h = headers(ws)
        id_idx = h.index(id_header) + 1
        row_num = None
        for i in range(2, ws.max_row + 1):
            if clean(ws.cell(i, id_idx).value) == values[id_header]:
                row_num = i
                break
        if row_num is None:
            row_num = ws.max_row + 1
        for key, value in values.items():
            set_cell(ws, row_num, key, value)

    upsert(
        source_ws,
        "来源ID",
        {
            "来源ID": "SRC-20260510-V08-RECALC",
            "来源类型": "系统重算",
            "省份": FUJIAN,
            "城市": "福建9城",
            "关联对象类型": "街道决策/KPI/字段核验",
            "关联对象名称": "V0.8街道机会重算",
            "资料标题": "周麻婆福建选址数据V0.8重算",
            "资料日期": TODAY,
            "采集日期": TODAY,
            "采集方式": "本地脚本重算",
            "文件/链接路径": str(Path(__file__).resolve()),
            "数据字段": "本店数、竞品数、五大竞品数、租金样本、POI、机会标签、字段缺口",
            "置信度影响": "修正导入后质量看板和街道机会判断",
            "可用结论": f"重新生成字段核验记录{matrix_rows_count}条，并生成本周街道核验任务",
            "限制说明": "街道匹配包含文本和半径近似，仍需地图/实地核验",
            "更新人": "Codex",
            "复核人": "待补",
            "备注": "不接高德API，不做外部抓取",
            "登记日期": TODAY,
            "登记人": "Codex",
            "复核状态": "待业务抽样复核",
        },
    )
    upsert(
        log_ws,
        "日志ID",
        {
            "日志ID": "LOG-20260510-V08-RECALC",
            "更新时间": TODAY,
            "更新类型": "质量重算/街道机会重算",
            "更新对象类型": "街道决策/KPI/质量看板/核验任务",
            "省份": FUJIAN,
            "城市": "福建9城",
            "对象名称": "V0.8福建街道机会",
            "更新前结论": "竞品和门店已导入，但字段缺口和街道机会尚未重算",
            "更新后结论": "完成街道空白机会、竞争验证、本店覆盖、租金/POI/照片缺口和本周核验任务重算",
            "变化原因": "吸收五大竞品和周麻婆门店后，需要把数据转成可执行选址动作",
            "更新人": "Codex",
            "复核人": "待补",
            "下一步动作": "优先实地核验福州高潜空白街道，并补租金/图片/平台招商截图",
            "备注": "低质量数据仍不得作为签约依据",
        },
    )


def main():
    workbook = workbook_path()
    backup = workbook.with_name(f"{workbook.stem}.bak-v08-recalc-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook.suffix}")
    shutil.copy2(workbook, backup)
    wb = load_workbook(workbook)
    update_street_decisions(wb)
    update_profiles(wb)
    rebuild_v08_tasks(wb)
    update_kpis(wb)
    matrix_rows = rebuild_field_matrix(wb)
    update_sources_and_log(wb, len(matrix_rows))
    wb.save(workbook)
    print(f"Workbook updated: {workbook}")
    print(f"Backup created: {backup}")
    print(f"Field matrix rows: {len(matrix_rows)}")
    print("V0.8 recalculation complete.")


if __name__ == "__main__":
    main()
