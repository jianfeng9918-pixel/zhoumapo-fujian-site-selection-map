from __future__ import annotations

import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "01-选址数据总表" / "周麻婆选址数据总表.xlsx"
TODAY = datetime.now().strftime("%Y-%m-%d")

FUJIAN = "福建"
FUJIAN_CITIES = {"福州", "厦门", "泉州", "漳州", "莆田", "三明", "南平", "龙岩", "宁德"}
CORE_COMPETITORS = {"小叫天", "醉得意", "四方桌", "大丰收", "姑奶奶"}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def to_float(value, default=0.0) -> float:
    try:
        return float(clean(value))
    except Exception:
        return default


def to_int(value, default=0) -> int:
    try:
        return int(float(clean(value)))
    except Exception:
        return default


def has_geo(row) -> bool:
    return clean(row.get("经度")) not in {"", "待补", "-"} and clean(row.get("纬度")) not in {"", "待补", "-"}


def workbook_path() -> Path:
    if WORKBOOK.exists():
        return WORKBOOK
    candidates = [path for path in ROOT.rglob("*.xlsx") if ".bak-" not in path.name and path.name != "brand_competitors.xlsx"]
    if not candidates:
        raise FileNotFoundError("未找到周麻婆选址数据总表")
    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def headers(ws) -> list[str]:
    return [clean(cell.value) for cell in ws[1]]


def records(ws) -> list[dict[str, str]]:
    h = headers(ws)
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row = {header: clean(ws.cell(row_num, col_num).value) for col_num, header in enumerate(h, start=1)}
        if any(row.values()):
            rows.append(row)
    return rows


def ensure_sheet(wb, name: str, h: list[str]):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    ws.append(h)
    return ws


def is_fujian(row) -> bool:
    return clean(row.get("省份")) == FUJIAN or clean(row.get("城市")) in FUJIAN_CITIES


def opportunity_class(tags: str, own_count: int) -> str:
    if "门店过密/谨慎" in tags:
        return "门店过密/谨慎"
    if "空白机会" in tags:
        return "空白机会"
    if "竞争验证强" in tags:
        return "竞争验证强"
    if own_count > 0 or "本店已覆盖" in tags:
        return "本店已覆盖"
    if "需租金核验" in tags:
        return "需租金核验"
    return "常规跟踪"


def build_district_metrics(wb):
    streets = [row for row in records(wb["街道决策表"]) if is_fujian(row)]
    stores = [row for row in records(wb["门店分布表"]) if is_fujian(row)]
    competitors = [row for row in records(wb["竞品门店库"]) if is_fujian(row)]
    tasks = [row for row in records(wb["核验任务表"]) if is_fujian(row) and clean(row.get("状态")) != "已完成"]
    field_rows = [row for row in records(wb["字段核验矩阵"]) if is_fujian(row) and clean(row.get("是否缺失")) == "是"]
    kpis = [row for row in records(wb["区县KPI表"]) if is_fujian(row)]

    result = []
    seen = set()
    for kpi in kpis:
        city = clean(kpi.get("城市"))
        district = clean(kpi.get("区县"))
        if not city or not district:
            continue
        key = (city, district)
        seen.add(key)
        ds = [row for row in streets if clean(row.get("城市")) == city and clean(row.get("区县")) == district]
        own = [row for row in stores if clean(row.get("城市")) == city and clean(row.get("区县")) == district]
        comp = [row for row in competitors if clean(row.get("城市")) == city and clean(row.get("区县")) == district]
        core = [row for row in comp if clean(row.get("竞品品牌")) in CORE_COMPETITORS]
        dist_tasks = [row for row in tasks if clean(row.get("城市")) == city and clean(row.get("区县")) == district]
        missing = [row for row in field_rows if clean(row.get("城市")) == city and clean(row.get("区县")) == district]
        high_blank = [row for row in ds if "空白机会" in clean(row.get("机会标签")) or "竞争验证强" in clean(row.get("机会标签"))]
        no_geo = [row for row in ds if not has_geo(row)]
        scores = [to_float(row.get("街道评分")) for row in ds if to_float(row.get("街道评分"))]
        qualities = [to_float(row.get("数据质量评分")) for row in ds if to_float(row.get("数据质量评分"))]
        result.append(
            {
                "指标ID": f"DMAP-{city}-{district}",
                "省份": FUJIAN,
                "城市": city,
                "区县": district,
                "推荐等级": clean(kpi.get("推荐等级")) or "待定",
                "综合评分": clean(kpi.get("综合评分")) or (round(sum(scores) / len(scores), 1) if scores else "待补"),
                "平均数据质量分": round(sum(qualities) / len(qualities), 1) if qualities else clean(kpi.get("平均数据质量分")) or "待补",
                "街道数": len(ds),
                "街道有坐标数": len(ds) - len(no_geo),
                "待定位街道数": len(no_geo),
                "周麻婆门店数": len(own),
                "五大竞品门店数": len(core),
                "全部竞品门店数": len(comp),
                "高潜空白街道数": len(high_blank),
                "P1任务数": sum(1 for row in dist_tasks if clean(row.get("优先级")) == "P1"),
                "待核验任务数": len(dist_tasks),
                "数据缺口数": len(missing),
                "主导机会类型": clean(Counter(opportunity_class(clean(row.get("机会标签")), to_int(row.get("周麻婆现有门店数"))) for row in ds).most_common(1)[0][0]) if ds else "待补",
                "数据缺口摘要": f"待定位街道{len(no_geo)}个；缺口字段{len(missing)}条；待核验任务{len(dist_tasks)}条",
                "下一步动作": "点击区县查看街道气泡和街道榜，优先核验高潜空白/竞争验证街道",
                "数据更新时间": TODAY,
            }
        )

    # Include any district that only appears in street decisions.
    for row in streets:
        key = (clean(row.get("城市")), clean(row.get("区县")))
        if key[0] and key[1] and key not in seen:
            city, district = key
            ds = [item for item in streets if clean(item.get("城市")) == city and clean(item.get("区县")) == district]
            result.append(
                {
                    "指标ID": f"DMAP-{city}-{district}",
                    "省份": FUJIAN,
                    "城市": city,
                    "区县": district,
                    "推荐等级": clean(row.get("推荐等级")) or "待定",
                    "综合评分": round(sum(to_float(item.get("街道评分")) for item in ds) / max(1, len(ds)), 1),
                    "平均数据质量分": round(sum(to_float(item.get("数据质量评分")) for item in ds) / max(1, len(ds)), 1),
                    "街道数": len(ds),
                    "街道有坐标数": sum(1 for item in ds if has_geo(item)),
                    "待定位街道数": sum(1 for item in ds if not has_geo(item)),
                    "周麻婆门店数": 0,
                    "五大竞品门店数": sum(to_int(item.get("五大竞品门店数")) for item in ds),
                    "全部竞品门店数": sum(to_int(item.get("友商/竞品门店数")) for item in ds),
                    "高潜空白街道数": sum(1 for item in ds if "空白机会" in clean(item.get("机会标签")) or "竞争验证强" in clean(item.get("机会标签"))),
                    "P1任务数": 0,
                    "待核验任务数": 0,
                    "数据缺口数": 0,
                    "主导机会类型": "街道样本",
                    "数据缺口摘要": "待补区县KPI",
                    "下一步动作": "补区县KPI并复核街道归属",
                    "数据更新时间": TODAY,
                }
            )
    return result


def build_street_points(wb):
    rows = [row for row in records(wb["街道决策表"]) if is_fujian(row)]
    result = []
    for row in rows:
        own = to_int(row.get("周麻婆现有门店数"))
        core = to_int(row.get("五大竞品门店数"))
        tags = clean(row.get("机会标签"))
        result.append(
            {
                "点ID": f"SMAP-{clean(row.get('决策ID'))}",
                "决策ID": clean(row.get("决策ID")),
                "省份": FUJIAN,
                "城市": clean(row.get("城市")),
                "区县": clean(row.get("区县")),
                "街道/片区": clean(row.get("街道/片区")),
                "关联商圈": clean(row.get("关联商圈")),
                "经度": clean(row.get("经度")) or "待补",
                "纬度": clean(row.get("纬度")) or "待补",
                "定位状态": "已定位" if has_geo(row) else "待定位",
                "机会标签": tags,
                "机会类型": opportunity_class(tags, own),
                "推荐等级": clean(row.get("推荐等级")) or "待定",
                "街道评分": clean(row.get("街道评分")),
                "数据质量评分": clean(row.get("数据质量评分")),
                "严格决策结论": clean(row.get("严格决策结论")),
                "适合店型": clean(row.get("适合店型")),
                "可开店容量": clean(row.get("可开店容量")),
                "周麻婆现有门店数": own,
                "五大竞品门店数": core,
                "友商/竞品门店数": to_int(row.get("友商/竞品门店数")),
                "租金样本数": to_int(row.get("租金样本数")),
                "数据缺口": clean(row.get("数据缺口")),
                "下一步核验动作": clean(row.get("下一步核验动作")),
                "来源等级": clean(row.get("来源等级")) or "L3",
                "气泡大小": max(5, min(18, round(to_float(row.get("街道评分")) / 7 + core / 2, 1))),
                "数据更新时间": TODAY,
            }
        )
    return result


def write_sheet(wb, name: str, rows: list[dict[str, str]], h: list[str]):
    ws = ensure_sheet(wb, name, h)
    for row in rows:
        ws.append([row.get(header, "") for header in h])


def main():
    workbook = workbook_path()
    backup = workbook.with_name(f"{workbook.stem}.bak-v09-district-map-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook.suffix}")
    shutil.copy2(workbook, backup)
    wb = load_workbook(workbook)
    district_rows = build_district_metrics(wb)
    street_rows = build_street_points(wb)
    write_sheet(
        wb,
        "区县地图指标",
        district_rows,
        ["指标ID", "省份", "城市", "区县", "推荐等级", "综合评分", "平均数据质量分", "街道数", "街道有坐标数", "待定位街道数", "周麻婆门店数", "五大竞品门店数", "全部竞品门店数", "高潜空白街道数", "P1任务数", "待核验任务数", "数据缺口数", "主导机会类型", "数据缺口摘要", "下一步动作", "数据更新时间"],
    )
    write_sheet(
        wb,
        "街道地图点",
        street_rows,
        ["点ID", "决策ID", "省份", "城市", "区县", "街道/片区", "关联商圈", "经度", "纬度", "定位状态", "机会标签", "机会类型", "推荐等级", "街道评分", "数据质量评分", "严格决策结论", "适合店型", "可开店容量", "周麻婆现有门店数", "五大竞品门店数", "友商/竞品门店数", "租金样本数", "数据缺口", "下一步核验动作", "来源等级", "气泡大小", "数据更新时间"],
    )
    wb.save(workbook)
    print(f"Workbook updated: {workbook}")
    print(f"Backup created: {backup}")
    print(f"District map metrics: {len(district_rows)}")
    print(f"Street map points: {len(street_rows)}")


if __name__ == "__main__":
    main()
